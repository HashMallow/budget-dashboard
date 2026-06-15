from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from django.db import transaction
from openpyxl import load_workbook

from marketing.importers.excel import ImportResult, discover_workbook, import_marketing_workbook, load_mapping
from marketing.models import Requester, SpendCategory, SubTeam, Vendor, normalize_name


@dataclass
class WorkbookLoadResult:
    import_result: ImportResult
    reference_result: ReferenceSeedResult


@dataclass
class ReferenceSeedCounter:
    created: int = 0
    updated: int = 0
    skipped: int = 0


@dataclass
class ReferenceSeedResult:
    workbook: Path
    dry_run: bool
    vendors: ReferenceSeedCounter = field(default_factory=ReferenceSeedCounter)
    categories: ReferenceSeedCounter = field(default_factory=ReferenceSeedCounter)
    sub_teams: ReferenceSeedCounter = field(default_factory=ReferenceSeedCounter)
    requesters: ReferenceSeedCounter = field(default_factory=ReferenceSeedCounter)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _upsert_by_normalized(
    model,
    name: str,
    *,
    source_sheet: str,
    counter: ReferenceSeedCounter,
    dry_run: bool,
    **extra,
):
    text = _cell_text(name)
    if not text:
        counter.skipped += 1
        return None
    normalized = normalize_name(text)
    if dry_run:
        exists = model.objects.filter(normalized_name=normalized).exists()
        if exists:
            counter.updated += 1
        else:
            counter.created += 1
        return None

    defaults = {"name": text, **extra}
    if hasattr(model, "source_sheet"):
        defaults["source_sheet"] = source_sheet

    obj, created = model.objects.get_or_create(normalized_name=normalized, defaults=defaults)
    if created:
        counter.created += 1
        return obj

    changed = False
    if obj.name != text:
        obj.name = text
        changed = True
    if hasattr(obj, "source_sheet") and source_sheet and obj.source_sheet != source_sheet:
        obj.source_sheet = source_sheet
        changed = True
    for key, value in extra.items():
        if getattr(obj, key) != value:
            setattr(obj, key, value)
            changed = True
    if changed:
        obj.save()
        counter.updated += 1
    else:
        counter.skipped += 1
    return obj


def seed_reference_data_from_workbook(
    workbook_path: str | Path | None = None,
    *,
    mapping_path: str | Path = "docs/discovery/column_mapping.yml",
    dry_run: bool = False,
) -> ReferenceSeedResult:
    path = discover_workbook(workbook_path)
    mapping = load_mapping(mapping_path)
    lookup = mapping.get("sheets", {}).get("lookup_data", {})
    sheet_name = lookup.get("actual_sheet_name", "Data")
    columns = lookup.get("columns", {})
    header_row = int(lookup.get("header_row", 1))

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Reference sheet '{sheet_name}' not found in workbook.")
    sheet = workbook[sheet_name]

    headers = {
        _cell_text(cell.value): idx for idx, cell in enumerate(sheet[header_row], start=1) if _cell_text(cell.value)
    }

    def col_index(key: str) -> int | None:
        header = columns.get(key)
        return headers.get(header) if header else None

    requester_col = col_index("requester_name")
    vendor_cols = [col_index("vendor_name"), col_index("unique_vendor_name")]
    category_cols = [col_index("title"), col_index("unique_title")]
    sub_team_col = col_index("sub_team")

    result = ReferenceSeedResult(workbook=path, dry_run=dry_run)

    def process_rows():
        for _row_number, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            row_values = list(row)

            def value_at(col: int | None, values=row_values) -> str:
                if col is None or col <= 0 or col > len(values):
                    return ""
                return _cell_text(values[col - 1])

            for col in vendor_cols:
                _upsert_by_normalized(
                    Vendor,
                    value_at(col),
                    source_sheet=sheet_name,
                    counter=result.vendors,
                    dry_run=dry_run,
                )
            for col in category_cols:
                _upsert_by_normalized(
                    SpendCategory,
                    value_at(col),
                    source_sheet=sheet_name,
                    counter=result.categories,
                    dry_run=dry_run,
                )
            _upsert_by_normalized(
                SubTeam,
                value_at(sub_team_col),
                source_sheet=sheet_name,
                counter=result.sub_teams,
                dry_run=dry_run,
            )
            _upsert_by_normalized(
                Requester,
                value_at(requester_col),
                source_sheet=sheet_name,
                counter=result.requesters,
                dry_run=dry_run,
            )

    if dry_run:
        process_rows()
    else:
        with transaction.atomic():
            process_rows()

    workbook.close()
    return result


def load_workbook_data(
    workbook_path: str | Path,
    *,
    mapping_path: str | Path = "docs/discovery/column_mapping.yml",
    dry_run: bool = False,
) -> WorkbookLoadResult:
    """Import invoices/budget lines and seed Data-sheet lookups (same as ``make load-data``)."""
    with transaction.atomic():
        import_result = import_marketing_workbook(
            workbook_path,
            mapping_path=mapping_path,
            dry_run=dry_run,
        )
        reference_result = seed_reference_data_from_workbook(
            workbook_path,
            mapping_path=mapping_path,
            dry_run=dry_run,
        )
    return WorkbookLoadResult(import_result=import_result, reference_result=reference_result)


def workbook_load_summary(load_result: WorkbookLoadResult) -> list[dict[str, int | str]]:
    imp = load_result.import_result
    ref = load_result.reference_result
    return [
        {
            "label": "Teams",
            "created": imp.teams.created,
            "updated": imp.teams.updated,
            "skipped": imp.teams.skipped,
        },
        {
            "label": "Vendors (invoices)",
            "created": imp.vendors.created,
            "updated": imp.vendors.updated,
            "skipped": imp.vendors.skipped,
        },
        {
            "label": "Campaigns",
            "created": imp.campaigns.created,
            "updated": imp.campaigns.updated,
            "skipped": imp.campaigns.skipped,
        },
        {
            "label": "Invoices",
            "created": imp.invoices.created,
            "updated": imp.invoices.updated,
            "skipped": imp.invoices.skipped,
        },
        {
            "label": "Budget",
            "created": imp.budget_lines.created,
            "updated": imp.budget_lines.updated,
            "skipped": imp.budget_lines.skipped,
        },
        {
            "label": "Vendors (lookup)",
            "created": ref.vendors.created,
            "updated": ref.vendors.updated,
            "skipped": ref.vendors.skipped,
        },
        {
            "label": "Categories",
            "created": ref.categories.created,
            "updated": ref.categories.updated,
            "skipped": ref.categories.skipped,
        },
        {
            "label": "Sub teams",
            "created": ref.sub_teams.created,
            "updated": ref.sub_teams.updated,
            "skipped": ref.sub_teams.skipped,
        },
        {
            "label": "Requesters",
            "created": ref.requesters.created,
            "updated": ref.requesters.updated,
            "skipped": ref.requesters.skipped,
        },
    ]
