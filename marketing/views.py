from __future__ import annotations

import random
from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from uuid import uuid4

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model, logout
from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST
from openpyxl import Workbook

from .analytics import (
    attention_invoices,
    decimal_sum,
    monthly_chart_data,
    monthly_spend_window_rows,
    overall_spend_pie,
    percent,
    team_chart_data,
    team_spend_rows,
    vendor_grouped_rows,
)
from .cost_buckets import exclude_pseudo_teams, team_spend_cost_buckets
from .exports.workbook import build_workbook_style_export
from .forms import (
    ExcelImportUploadForm,
    InvoiceAttachmentForm,
    InvoiceForm,
    InvoiceStatusForm,
    UserAccessCreateForm,
    user_can_create_invoice,
)
from .importers.excel import ImportResult, import_marketing_workbook
from .jalali import JALALI_MONTHS, gregorian_to_jalali, jalali_year_bounds, today_jalali
from .models import (
    BudgetLine,
    CostBucket,
    Invoice,
    PaymentStage,
    Team,
    UserTeamAccess,
)
from .permissions import (
    can_edit_invoice,
    can_export,
    filter_budget_lines_for_user,
    filter_invoices_for_user,
    filter_teams_for_user,
    get_user_scope,
    user_has_admin_access,
)
from .reports.pdf import build_dashboard_summary_pdf
from .table_sort import SortState, apply_ordering, parse_sort, sort_rows

User = get_user_model()
ZERO = Decimal("0")

# Small set of gradient palettes + glyphs used to build a random tab logo (favicon).
_FAVICON_PALETTES = [
    ("#0f766e", "#14b8a6"),
    ("#7c3aed", "#a855f7"),
    ("#2563eb", "#06b6d4"),
    ("#dc2626", "#f97316"),
    ("#db2777", "#f43f5e"),
    ("#ca8a04", "#facc15"),
    ("#059669", "#34d399"),
    ("#4f46e5", "#818cf8"),
]
_FAVICON_GLYPHS = ["📊", "📈", "💸", "📉", "🧾", "💹", "🎯", "📌", "🪙", "📣"]


def favicon_svg(request):
    """Return a randomly themed SVG app icon so the browser tab shows a fun logo, not a blank box."""
    color_start, color_end = random.choice(_FAVICON_PALETTES)
    glyph = random.choice(_FAVICON_GLYPHS)
    svg = (
        '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64">'
        '<defs><linearGradient id="g" x1="0" y1="0" x2="1" y2="1">'
        f'<stop offset="0" stop-color="{color_start}"/><stop offset="1" stop-color="{color_end}"/>'
        "</linearGradient></defs>"
        '<rect width="64" height="64" rx="14" fill="url(#g)"/>'
        '<text x="32" y="33" font-size="36" text-anchor="middle" '
        f'dominant-baseline="central">{glyph}</text>'
        "</svg>"
    )
    response = HttpResponse(svg, content_type="image/svg+xml")
    # Re-roll the logo on every visit instead of caching a single icon.
    response["Cache-Control"] = "no-store, max-age=0"
    return response

INVOICE_SORT_FIELDS = {
    "number": "invoice_number",
    "vendor": "vendor__name",
    "team": "team__name",
    "category": "category",
    "date": "invoice_date",
    "amount": "amount",
    "stage": "payment_stage",
    "days": "stage_changed_at",
}
INVOICE_SORT_DEFAULTS = {
    "number": "asc",
    "vendor": "asc",
    "team": "asc",
    "category": "asc",
    "date": "desc",
    "amount": "desc",
    "stage": "asc",
    "days": "desc",
}

VENDOR_SORT_KEYS = {
    "vendor": lambda row: row["vendor"].name.lower(),
    "invoices": lambda row: row["invoice_count"],
    "amount": lambda row: row["total"],
}
VENDOR_SORT_DEFAULTS = {"vendor": "asc", "invoices": "desc", "amount": "desc"}

CAMPAIGN_SORT_FIELDS = {
    "campaign": "campaign__name",
    "year": "campaign__year",
    "team": "campaign__team__name",
    "invoices": "invoice_count",
    "amount": "total",
}
CAMPAIGN_SORT_DEFAULTS = {
    "campaign": "asc",
    "year": "desc",
    "team": "asc",
    "invoices": "desc",
    "amount": "desc",
}

TEAM_SORT_KEYS = {
    "team": lambda row: row["team"].name.lower(),
    "invoices": lambda row: row["invoice_count"],
    "amount": lambda row: row["total_spend"],
}
TEAM_SORT_DEFAULTS = {"team": "asc", "invoices": "desc", "amount": "desc"}

BUDGET_SORT_FIELDS = {
    "year": "year",
    "month": "month",
    "team": "team__name",
    "campaign": "campaign__name",
    "category": "category",
    "amount": "planned_amount",
}
BUDGET_SORT_DEFAULTS = {
    "year": "desc",
    "month": "asc",
    "team": "asc",
    "campaign": "asc",
    "category": "asc",
    "amount": "desc",
}
BUDGET_PIVOT_SORT_KEYS = {
    "team": lambda row: row["team"].lower(),
    "category": lambda row: row["category"].lower(),
    "total": lambda row: row["total"],
}


def get_ui_lang(request) -> str:
    ui_lang = request.session.get("ui_lang", "en")
    return ui_lang if ui_lang in {"fa", "en"} else "en"


def get_months(request) -> list[tuple[int, str]]:
    """Persian (Jalali) month labels, localized to the active UI language."""
    if request.session.get("ui_lang", "en") == "fa":
        return [(number, persian) for number, persian, _latin in JALALI_MONTHS]
    return [(number, latin) for number, _persian, latin in JALALI_MONTHS]


def monthly_trend_window(selected_year: str) -> tuple[int, int, int]:
    """Window (end_year, end_month, count) for the monthly trend chart.

    Avoids showing future months: the current Jalali year stops at the current month, and the
    default ("all years") view is a trailing 12 months. A specific past year shows its 12 months.
    """
    current_year, current_month, _day = today_jalali()
    if selected_year.isdigit():
        year = int(selected_year)
        end_month = current_month if year == current_year else 12
        return year, end_month, end_month
    return current_year, current_month, 12


def filter_by_jalali_year(queryset, year: str):
    """Filter invoices by a Jalali year using the matching Gregorian date range."""
    if year and year.isdigit():
        start, end = jalali_year_bounds(int(year))
        return queryset.filter(invoice_date__range=(start, end))
    return queryset


def distinct_jalali_years(queryset) -> list[int]:
    # Small datasets: convert each invoice date. Revisit with a stored Jalali year
    # column if invoice volume grows large.
    years = {
        gregorian_to_jalali(value.year, value.month, value.day)[0]
        for value in queryset.values_list("invoice_date", flat=True)
        if value
    }
    return sorted(years, reverse=True)


def forbidden(message: str = "You are not allowed to perform this action.") -> HttpResponseForbidden:
    return HttpResponseForbidden(message)


def visible_invoice_queryset(request):
    queryset = Invoice.objects.select_related("vendor", "team", "campaign").order_by("-invoice_date", "-id")
    return filter_invoices_for_user(queryset, request.user)


def visible_team_queryset(request):
    teams = exclude_pseudo_teams(Team.objects.filter(is_active=True))
    return filter_teams_for_user(teams, request.user).order_by("name")


def filter_invoice_queryset(request, queryset):
    filters = {
        "q": request.GET.get("q", "").strip(),
        "year": request.GET.get("year", "").strip(),
        "team": request.GET.get("team", "").strip(),
        "stage": request.GET.get("stage", "").strip(),
        "bucket": request.GET.get("bucket", "").strip(),
    }

    if filters["q"]:
        queryset = queryset.filter(
            Q(invoice_number__icontains=filters["q"])
            | Q(vendor__name__icontains=filters["q"])
            | Q(campaign__name__icontains=filters["q"])
            | Q(category__icontains=filters["q"])
            | Q(description__icontains=filters["q"])
        )
    if filters["year"].isdigit():
        queryset = filter_by_jalali_year(queryset, filters["year"])
    if filters["team"].isdigit():
        queryset = queryset.filter(team_id=int(filters["team"]))
    if filters["stage"]:
        queryset = queryset.filter(payment_stage=filters["stage"])
    if filters["bucket"]:
        queryset = queryset.filter(cost_bucket=filters["bucket"])
    return queryset, filters


def result_summary(result: ImportResult) -> list[dict[str, int | str]]:
    return [
        {
            "label": "Teams",
            "created": result.teams.created,
            "updated": result.teams.updated,
            "skipped": result.teams.skipped,
        },
        {
            "label": "Vendors",
            "created": result.vendors.created,
            "updated": result.vendors.updated,
            "skipped": result.vendors.skipped,
        },
        {
            "label": "Campaigns",
            "created": result.campaigns.created,
            "updated": result.campaigns.updated,
            "skipped": result.campaigns.skipped,
        },
        {
            "label": "Invoices",
            "created": result.invoices.created,
            "updated": result.invoices.updated,
            "skipped": result.invoices.skipped,
        },
        {
            "label": "Budget",
            "created": result.budget_lines.created,
            "updated": result.budget_lines.updated,
            "skipped": result.budget_lines.skipped,
        },
    ]


@login_required
def dashboard(request):
    base_queryset = visible_invoice_queryset(request)
    years = distinct_jalali_years(base_queryset)
    selected_year = request.GET.get("year", "").strip()
    selected_team = request.GET.get("team", "").strip()

    invoices = base_queryset
    if selected_year.isdigit():
        invoices = filter_by_jalali_year(invoices, selected_year)
    if selected_team.isdigit():
        invoices = invoices.filter(team_id=int(selected_team))

    total_spend = decimal_sum(invoices)
    paid_spend = decimal_sum(invoices.filter(payment_stage=PaymentStage.PAID))
    invoice_count = invoices.count()
    referral_total = decimal_sum(invoices.filter(cost_bucket=CostBucket.REFERRAL))
    sms_total = decimal_sum(invoices.filter(cost_bucket=CostBucket.SMS))

    end_year, end_month, window_count = monthly_trend_window(selected_year)
    monthly_rows = monthly_spend_window_rows(
        invoices,
        end_year=end_year,
        end_month=end_month,
        count=window_count,
        ui_lang=get_ui_lang(request),
    )
    monthly_chart = monthly_chart_data(monthly_rows)

    team_total_rows = team_spend_rows(invoices)
    team_chart = team_chart_data(team_total_rows, get_ui_lang(request))

    vendor_rows = list(
        invoices.values("vendor_id", "vendor__name")
        .annotate(total=Sum("amount"), invoice_count=Count("id"))
        .order_by("-total")[:8]
    )

    campaign_rows = list(
        invoices.filter(campaign__isnull=False)
        .values("campaign_id", "campaign__name", "campaign__year")
        .annotate(total=Sum("amount"), invoice_count=Count("id"))
        .order_by("-total")[:8]
    )

    stage_rows = list(
        invoices.values("payment_stage")
        .annotate(invoice_count=Count("id"), total=Sum("amount"))
        .order_by("-invoice_count")
    )

    attention = attention_invoices(invoices)

    ui_lang = get_ui_lang(request)
    spend_pie = overall_spend_pie(team_total_rows, referral_total, sms_total, ui_lang)

    context = {
        "scope": get_user_scope(request.user),
        "years": years,
        "selected_year": selected_year,
        "selected_team": selected_team,
        "teams": visible_team_queryset(request),
        "total_spend": total_spend,
        "paid_spend": paid_spend,
        "invoice_count": invoice_count,
        "referral_total": referral_total,
        "sms_total": sms_total,
        "monthly_rows": monthly_rows,
        "team_total_rows": team_total_rows,
        "vendor_rows": vendor_rows,
        "campaign_rows": campaign_rows,
        "stage_rows": stage_rows,
        "attention_invoices": attention,
        "spend_pie": spend_pie,
        "spend_pie_has_data": bool(spend_pie["values"]),
        "monthly_chart": monthly_chart,
        "monthly_chart_has_data": any(value for value in monthly_chart["values"]),
        "team_chart": team_chart,
        "team_chart_has_data": bool(team_chart["values"]),
        "can_create_invoice": user_can_create_invoice(request.user),
        "can_export_data": can_export(request.user),
    }
    return render(request, "marketing/dashboard.html", context)


@login_required
def invoice_list(request):
    queryset, filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    sort = parse_sort(
        request,
        allowed=INVOICE_SORT_FIELDS,
        default_field="date",
        default_dir="desc",
        default_dirs=INVOICE_SORT_DEFAULTS,
    )
    queryset = apply_ordering(
        queryset,
        sort,
        fields=INVOICE_SORT_FIELDS,
        default_field="date",
        inverted={"days"},
        tiebreaker="-id",
    )
    paginator = Paginator(queryset, 25)
    page_obj = paginator.get_page(request.GET.get("page"))
    context = {
        "page_obj": page_obj,
        "filters": filters,
        "teams": visible_team_queryset(request),
        "years": distinct_jalali_years(visible_invoice_queryset(request)),
        "payment_stages": PaymentStage.choices,
        "cost_buckets": CostBucket.choices,
        "can_create_invoice": user_can_create_invoice(request.user),
        "can_export_data": can_export(request.user),
        "table_sort": sort,
        "table_sort_defaults": INVOICE_SORT_DEFAULTS,
    }
    return render(request, "marketing/invoices/list.html", context)


@login_required
def invoice_detail(request, pk: int):
    invoice = get_object_or_404(visible_invoice_queryset(request), pk=pk)
    status_form = InvoiceStatusForm(invoice=invoice, ui_lang=get_ui_lang(request))
    attachment_form = InvoiceAttachmentForm(user=request.user, invoice=invoice, ui_lang=get_ui_lang(request))
    context = {
        "invoice": invoice,
        "status_form": status_form,
        "attachment_form": attachment_form,
        "can_edit": can_edit_invoice(request.user, invoice),
        "can_upload": attachment_form.has_allowed_types,
    }
    return render(request, "marketing/invoices/detail.html", context)


@login_required
def invoice_create(request):
    if not user_can_create_invoice(request.user):
        return forbidden()
    if request.method == "POST":
        form = InvoiceForm(request.POST, user=request.user, ui_lang=get_ui_lang(request))
        if form.is_valid():
            invoice = form.save()
            messages.success(request, "Invoice saved.")
            return redirect("marketing:invoice_detail", pk=invoice.pk)
    else:
        form = InvoiceForm(user=request.user, ui_lang=get_ui_lang(request))
    return render(request, "marketing/invoices/form.html", {"form": form, "mode": "create"})


@login_required
def invoice_edit(request, pk: int):
    invoice = get_object_or_404(visible_invoice_queryset(request), pk=pk)
    if not can_edit_invoice(request.user, invoice):
        return forbidden()
    if request.method == "POST":
        form = InvoiceForm(request.POST, user=request.user, instance=invoice, ui_lang=get_ui_lang(request))
        if form.is_valid():
            invoice = form.save()
            messages.success(request, "Invoice updated.")
            return redirect("marketing:invoice_detail", pk=invoice.pk)
    else:
        form = InvoiceForm(user=request.user, instance=invoice, ui_lang=get_ui_lang(request))
    return render(request, "marketing/invoices/form.html", {"form": form, "invoice": invoice, "mode": "edit"})


@login_required
@require_POST
def invoice_stage_update(request, pk: int):
    invoice = get_object_or_404(visible_invoice_queryset(request), pk=pk)
    if not can_edit_invoice(request.user, invoice):
        return forbidden()
    form = InvoiceStatusForm(request.POST, invoice=invoice, ui_lang=get_ui_lang(request))
    if form.is_valid():
        invoice.set_payment_stage(
            form.cleaned_data["payment_stage"],
            changed_by=request.user,
            note=form.cleaned_data.get("note", ""),
        )
        messages.success(request, "Payment stage updated.")
    else:
        messages.error(request, "Invalid payment stage.")
    return redirect("marketing:invoice_detail", pk=invoice.pk)


@login_required
@require_POST
def invoice_attachment_upload(request, pk: int):
    invoice = get_object_or_404(visible_invoice_queryset(request), pk=pk)
    form = InvoiceAttachmentForm(
        request.POST,
        request.FILES,
        user=request.user,
        invoice=invoice,
        ui_lang=get_ui_lang(request),
    )
    if not form.has_allowed_types:
        return forbidden("You are not allowed to upload files for this invoice.")
    if form.is_valid():
        attachment = form.save(commit=False)
        attachment.invoice = invoice
        attachment.uploaded_by = request.user
        attachment.save()
        messages.success(request, "File uploaded.")
    else:
        messages.error(request, "Upload failed. Check the file type or your permissions.")
    return redirect("marketing:invoice_detail", pk=invoice.pk)


@login_required
def team_list(request):
    sort = parse_sort(
        request,
        allowed=TEAM_SORT_KEYS,
        default_field="amount",
        default_dir="desc",
        default_dirs=TEAM_SORT_DEFAULTS,
    )
    teams = visible_team_queryset(request)
    team_summaries = []
    for team in teams:
        invoices = visible_invoice_queryset(request).filter(
            team=team,
            cost_bucket__in=team_spend_cost_buckets(team),
        )
        team_summaries.append({
            "team": team,
            "total_spend": decimal_sum(invoices),
            "invoice_count": invoices.count(),
        })
    team_summaries = sort_rows(
        team_summaries,
        sort,
        keys=TEAM_SORT_KEYS,
        default_field="amount",
    )
    return render(
        request,
        "marketing/teams/list.html",
        {
            "team_summaries": team_summaries,
            "table_sort": sort,
            "table_sort_defaults": TEAM_SORT_DEFAULTS,
        },
    )


@login_required
def team_dashboard(request, pk: int):
    team = get_object_or_404(visible_team_queryset(request), pk=pk)
    selected_year = request.GET.get("year", "").strip()

    invoices = visible_invoice_queryset(request).filter(
        team=team,
        cost_bucket__in=team_spend_cost_buckets(team),
    )
    if selected_year.isdigit():
        invoices = filter_by_jalali_year(invoices, selected_year)

    total_spend = decimal_sum(invoices)
    invoice_count = invoices.count()
    end_year, end_month, window_count = monthly_trend_window(selected_year)
    monthly_rows = monthly_spend_window_rows(
        invoices,
        end_year=end_year,
        end_month=end_month,
        count=window_count,
        ui_lang=get_ui_lang(request),
    )
    monthly_chart = monthly_chart_data(monthly_rows)
    vendor_rows = vendor_grouped_rows(invoices)
    campaign_rows = list(
        invoices.filter(campaign__isnull=False)
        .values("campaign_id", "campaign__name", "campaign__year")
        .annotate(total=Sum("amount"), invoice_count=Count("id"))
        .order_by("-total")
    )
    attention = attention_invoices(invoices)
    budget_total = decimal_sum(
        filter_budget_lines_for_user(BudgetLine.objects.filter(team=team), request.user),
        "planned_amount",
    )

    export_query = f"team={team.id}"
    if selected_year.isdigit():
        export_query += f"&year={selected_year}"

    context = {
        "team": team,
        "years": distinct_jalali_years(visible_invoice_queryset(request).filter(team=team)),
        "selected_year": selected_year,
        "total_spend": total_spend,
        "budget_total": budget_total,
        "invoice_count": invoice_count,
        "monthly_rows": monthly_rows,
        "monthly_chart": monthly_chart,
        "monthly_chart_has_data": any(value for value in monthly_chart["values"]),
        "vendor_rows": vendor_rows,
        "campaign_rows": campaign_rows,
        "attention_invoices": attention,
        "can_create_invoice": user_can_create_invoice(request.user),
        "can_export_data": can_export(request.user),
        "export_query": export_query,
    }
    return render(request, "marketing/teams/dashboard.html", context)


@login_required
def vendor_report(request):
    queryset, filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    sort = parse_sort(
        request,
        allowed=VENDOR_SORT_KEYS,
        default_field="amount",
        default_dir="desc",
        default_dirs=VENDOR_SORT_DEFAULTS,
    )
    vendor_rows = sort_rows(
        vendor_grouped_rows(queryset),
        sort,
        keys=VENDOR_SORT_KEYS,
        default_field="amount",
    )

    context = {
        "vendor_rows": vendor_rows,
        "filters": filters,
        "teams": visible_team_queryset(request),
        "payment_stages": PaymentStage.choices,
        "cost_buckets": CostBucket.choices,
        "can_export_data": can_export(request.user),
        "table_sort": sort,
        "table_sort_defaults": VENDOR_SORT_DEFAULTS,
    }
    return render(request, "marketing/vendors/report.html", context)


@login_required
def campaign_report(request):
    months = get_months(request)
    queryset, filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    sort = parse_sort(
        request,
        allowed=CAMPAIGN_SORT_FIELDS,
        default_field="amount",
        default_dir="desc",
        default_dirs=CAMPAIGN_SORT_DEFAULTS,
    )
    campaign_qs = (
        queryset.filter(campaign__isnull=False)
        .values("campaign_id", "campaign__name", "campaign__year", "campaign__team__name")
        .annotate(total=Sum("amount"), invoice_count=Count("id"))
    )
    campaign_qs = apply_ordering(
        campaign_qs,
        sort,
        fields=CAMPAIGN_SORT_FIELDS,
        default_field="amount",
        tiebreaker="campaign__name",
    )
    campaign_totals = list(campaign_qs)
    max_total = max((row["total"] or ZERO for row in campaign_totals), default=ZERO)
    for row in campaign_totals:
        row["percent"] = percent(row["total"] or ZERO, max_total)

    monthly_campaigns = defaultdict(lambda: {month: ZERO for month, _label in months})
    for row in (
        queryset.filter(campaign__isnull=False)
        .values("campaign__name", "invoice_date")
        .annotate(total=Sum("amount"))
    ):
        invoice_date = row["invoice_date"]
        if not invoice_date:
            continue
        jalali_month = gregorian_to_jalali(invoice_date.year, invoice_date.month, invoice_date.day)[1]
        monthly_campaigns[row["campaign__name"]][jalali_month] += row["total"] or ZERO

    context = {
        "campaign_rows": campaign_totals,
        "monthly_campaigns": dict(monthly_campaigns),
        "months": months,
        "filters": filters,
        "teams": visible_team_queryset(request),
        "years": distinct_jalali_years(visible_invoice_queryset(request)),
        "payment_stages": PaymentStage.choices,
        "cost_buckets": CostBucket.choices,
        "can_export_data": can_export(request.user),
        "table_sort": sort,
        "table_sort_defaults": CAMPAIGN_SORT_DEFAULTS,
    }
    return render(request, "marketing/campaigns/report.html", context)


@login_required
def budget_list(request):
    months = get_months(request)
    sort = parse_sort(
        request,
        allowed={*BUDGET_SORT_FIELDS, "total"},
        default_field="year",
        default_dir="desc",
        default_dirs={**BUDGET_SORT_DEFAULTS, "total": "desc"},
    )
    queryset = filter_budget_lines_for_user(
        BudgetLine.objects.select_related("team", "campaign"),
        request.user,
    )
    year = request.GET.get("year", "").strip()
    team = request.GET.get("team", "").strip()
    q = request.GET.get("q", "").strip()
    if year.isdigit():
        queryset = queryset.filter(year=int(year))
    if team.isdigit():
        queryset = queryset.filter(team_id=int(team))
    if q:
        queryset = queryset.filter(Q(category__icontains=q) | Q(team__name__icontains=q))

    db_sort = sort if sort.field in BUDGET_SORT_FIELDS else SortState("year", sort.direction)
    queryset = apply_ordering(
        queryset,
        db_sort,
        fields=BUDGET_SORT_FIELDS,
        default_field="year",
        tiebreaker="id",
    )

    years = (
        filter_budget_lines_for_user(BudgetLine.objects.all(), request.user)
        .order_by("-year")
        .values_list("year", flat=True)
        .distinct()
    )
    pivot = {}
    for line in queryset:
        key = (line.team.name if line.team else "No team", line.category)
        if key not in pivot:
            pivot[key] = {
                "team": key[0],
                "category": key[1],
                "months": {month_number: ZERO for month_number, _label in months},
                "total": ZERO,
            }
        if line.month:
            pivot[key]["months"][line.month] += line.planned_amount or ZERO
        pivot[key]["total"] += line.planned_amount or ZERO

    pivot_sort = SortState(
        sort.field if sort.field in BUDGET_PIVOT_SORT_KEYS else "team",
        sort.direction,
    )
    pivot_rows = sort_rows(
        list(pivot.values()),
        pivot_sort,
        keys=BUDGET_PIVOT_SORT_KEYS,
        default_field="team",
    )

    paginator = Paginator(queryset, 50)
    context = {
        "page_obj": paginator.get_page(request.GET.get("page")),
        "pivot_rows": pivot_rows,
        "months": months,
        "years": years,
        "teams": visible_team_queryset(request),
        "filters": {"year": year, "team": team, "q": q},
        "table_sort": sort,
        "table_sort_defaults": {**BUDGET_SORT_DEFAULTS, "total": "desc"},
    }
    return render(request, "marketing/budgets/list.html", context)


@login_required
def import_workbook(request):
    if not user_has_admin_access(request.user):
        return forbidden()

    form = ExcelImportUploadForm(ui_lang=get_ui_lang(request))
    result = None
    summary = None
    pending_path = request.session.get("pending_import_path")

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "confirm":
            if not pending_path:
                messages.error(request, "There is no file ready to import.")
            else:
                result = import_marketing_workbook(pending_path, dry_run=False)
                summary = result_summary(result)
                messages.success(request, "Excel data imported into the database.")
                request.session.pop("pending_import_path", None)
                pending_path = None
        else:
            form = ExcelImportUploadForm(request.POST, request.FILES, ui_lang=get_ui_lang(request))
            if form.is_valid():
                storage = FileSystemStorage(location=Path(settings.MEDIA_ROOT) / "imports")
                filename = f"{uuid4().hex}.xlsx"
                stored_name = storage.save(filename, form.cleaned_data["workbook"])
                workbook_path = storage.path(stored_name)
                result = import_marketing_workbook(workbook_path, dry_run=True)
                summary = result_summary(result)
                request.session["pending_import_path"] = workbook_path
                pending_path = workbook_path
                messages.info(request, "Dry-run complete. If the result looks correct, confirm the import.")

    context = {
        "form": form,
        "result": result,
        "summary": summary,
        "pending_path": pending_path,
        "skipped_preview": result.skipped_rows[:20] if result else [],
        "can_export_data": can_export(request.user),
    }
    return render(request, "marketing/imports/upload.html", context)


@login_required
def user_access(request):
    if not user_has_admin_access(request.user):
        return forbidden()

    form = UserAccessCreateForm(user=request.user, ui_lang=get_ui_lang(request))
    if request.method == "POST":
        action = request.POST.get("action")
        if action in {"deactivate", "activate"}:
            target = get_object_or_404(User, pk=request.POST.get("user_id"))
            if target == request.user and action == "deactivate":
                messages.error(request, "You cannot deactivate your own account.")
            else:
                target.is_active = action == "activate"
                target.save(update_fields=["is_active"])
                UserTeamAccess.objects.filter(user=target).update(is_active=target.is_active)
                messages.success(request, "User status updated.")
            return redirect("marketing:user_access")

        form = UserAccessCreateForm(request.POST, user=request.user, ui_lang=get_ui_lang(request))
        if form.is_valid():
            form.save()
            messages.success(request, "New user created.")
            return redirect("marketing:user_access")

    users = User.objects.prefetch_related("groups", "team_access__team").order_by("username")
    return render(request, "marketing/users/access.html", {"form": form, "users": users})


@login_required
def export_invoices_excel(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    queryset, _filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoices"
    sheet.append([
        "Invoice number",
        "Vendor",
        "Team",
        "Campaign",
        "Category",
        "Cost bucket",
        "Invoice date",
        "Amount",
        "Currency",
        "Payment stage",
        "Days in stage",
    ])
    for invoice in queryset:
        sheet.append([
            invoice.invoice_number,
            invoice.vendor.name,
            invoice.team.name if invoice.team else "",
            invoice.campaign.name if invoice.campaign else "",
            invoice.category,
            invoice.cost_bucket,
            invoice.invoice_date.isoformat(),
            invoice.amount,
            invoice.currency,
            invoice.payment_stage,
            invoice.days_in_current_stage,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="marketing-invoices.xlsx"'
    workbook.save(response)
    return response


@login_required
def export_vendors_excel(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    queryset, _filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    vendor_rows = vendor_grouped_rows(queryset)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Vendors"
    sheet.append(["Vendor", "Invoice count", "Invoice numbers", "Payment stages", "Total spend"])
    for row in vendor_rows:
        sheet.append([
            row["vendor"].name,
            row["invoice_count"],
            ", ".join(row["invoice_numbers"]),
            ", ".join(row["stages"]),
            row["total"],
        ])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="marketing-vendors.xlsx"'
    workbook.save(response)
    return response


@login_required
def export_campaigns_excel(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    queryset, _filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    campaign_rows = (
        queryset.filter(campaign__isnull=False)
        .values("campaign__name", "campaign__year", "campaign__team__name")
        .annotate(total=Sum("amount"), invoice_count=Count("id"))
        .order_by("-total")
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Campaigns"
    sheet.append(["Campaign", "Year", "Team", "Invoice count", "Total spend"])
    for row in campaign_rows:
        sheet.append([
            row["campaign__name"],
            row["campaign__year"],
            row["campaign__team__name"] or "",
            row["invoice_count"],
            row["total"],
        ])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="marketing-campaigns.xlsx"'
    workbook.save(response)
    return response


@login_required
def export_workbook_excel(request):
    """Export an Excel file shaped like the source workbook (same sheet names and layout).

    This recreates the familiar workbook structure from the database — the invoice sheet,
    the wide monthly Budget projection/Actual sheet, the Market Live Spending summary, and
    the Data lookup lists — scoped to the data the current user is allowed to see.
    """
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    invoices, filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    budget_lines = filter_budget_lines_for_user(BudgetLine.objects.all(), request.user)
    if filters["year"].isdigit():
        budget_lines = budget_lines.filter(year=int(filters["year"]))
    if filters["team"].isdigit():
        budget_lines = budget_lines.filter(team_id=int(filters["team"]))

    workbook = build_workbook_style_export(invoices, budget_lines)
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="marketing-workbook.xlsx"'
    workbook.save(response)
    return response


@login_required
def dashboard_report_pdf(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    queryset, filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    vendor_rows = list(
        queryset.values("vendor__name")
        .annotate(total=Sum("amount"), invoice_count=Count("id"))
        .order_by("-total")[:15]
    )
    stage_rows = list(
        queryset.values("payment_stage").annotate(invoice_count=Count("id")).order_by("payment_stage")
    )
    pdf_bytes = build_dashboard_summary_pdf(
        title="Marketing spend summary",
        generated_at=timezone.now(),
        total_spend=decimal_sum(queryset),
        invoice_count=queryset.count(),
        vendor_rows=vendor_rows,
        stage_rows=stage_rows,
        filters=filters,
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="marketing-dashboard-summary.pdf"'
    return response


@login_required
def invoice_report_print(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to view reports.")
    queryset, filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    vendor_rows = list(
        queryset.values("vendor__name")
        .annotate(total=Sum("amount"), invoice_count=Count("id"))
        .order_by("-total")[:30]
    )
    context = {
        "generated_at": timezone.now(),
        "filters": filters,
        "total_spend": decimal_sum(queryset),
        "invoice_count": queryset.count(),
        "vendor_rows": vendor_rows,
        "stage_rows": queryset.values("payment_stage").annotate(invoice_count=Count("id")).order_by("payment_stage"),
    }
    return render(request, "marketing/print/invoice_report.html", context)


@require_POST
def set_display_preferences(request):
    ui_lang = request.POST.get("ui_lang")
    number_locale = request.POST.get("number_locale")
    money_format = request.POST.get("money_format")
    currency_unit = request.POST.get("currency_unit")
    theme = request.POST.get("theme")
    if ui_lang in {"fa", "en"}:
        request.session["ui_lang"] = ui_lang
    if number_locale in {"fa", "en"}:
        request.session["number_locale"] = number_locale
    if money_format in {"full", "compact"}:
        request.session["money_format"] = money_format
    if currency_unit in {"rial", "toman"}:
        request.session["currency_unit"] = currency_unit
    if theme in {"light", "dark"}:
        request.session["theme"] = theme

    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or "/"
    if not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        next_url = "/"
    return redirect(next_url)


@require_POST
def logout_view(request):
    logout(request)
    return redirect("marketing:login")
