from __future__ import annotations

from decimal import Decimal

from django.contrib.auth import get_user_model
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from urllib.parse import urlencode

from ..analytics import (
    decimal_sum,
    vendor_grouped_rows,
)
from .core import (
    distinct_business_sections,
    distinct_jalali_years,
    filter_contract_queryset,
    filter_invoice_queryset,
    forbidden,
    get_months,
    get_ui_lang,
    pdf_locale_for_request,
    visible_contract_queryset,
    visible_invoice_queryset,
    visible_team_queryset,
)
from ..exports.workbook import build_workbook_style_export
from ..models import (
    BudgetLine,
    Vendor,
)
from ..permissions import (
    can_export,
    filter_budget_lines_for_user,
)
from ..reports.pdf import (
    build_campaign_report_pdf,
    build_contract_report_pdf,
    build_dashboard_summary_pdf,
    build_vendor_report_pdf,
)
from ..translations import translate

User = get_user_model()
ZERO = Decimal("0")
_FAVICON_SVG = (
    '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64">'
    '<defs><linearGradient id="g" x1="0" y1="0" x2="1" y2="1">'
    '<stop offset="0" stop-color="#ca8a04"/><stop offset="1" stop-color="#facc15"/>'
    "</linearGradient></defs>"
    '<rect width="64" height="64" rx="14" fill="url(#g)"/>'
    '<rect x="18" y="14" width="28" height="36" rx="3" fill="#fff"/>'
    '<circle cx="24" cy="20" r="3" fill="#2563eb"/>'
    '<line x1="22" y1="26" x2="42" y2="26" stroke="#2563eb" stroke-width="2" stroke-linecap="round"/>'
    '<line x1="22" y1="32" x2="42" y2="32" stroke="#2563eb" stroke-width="2" stroke-linecap="round"/>'
    '<line x1="22" y1="38" x2="38" y2="38" stroke="#2563eb" stroke-width="2" stroke-linecap="round"/>'
    "</svg>"
)
INVOICE_SORT_FIELDS = {
    "number": "invoice_number",
    "vendor": "vendor__name",
    "team": "team__name",
    "category": "category",
    "date": "invoice_date",
    "amount": "amount",
    "stage": "payment_stage",
    "days": "stage_changed_at",
}
INVOICE_SORT_DEFAULTS = {
    "number": "asc",
    "vendor": "asc",
    "team": "asc",
    "category": "asc",
    "date": "desc",
    "amount": "desc",
    "stage": "asc",
    "days": "desc",
}
VENDOR_SORT_KEYS = {
    "vendor": lambda row: row["vendor"].name.lower(),
    "invoices": lambda row: row["invoice_count"],
    "amount": lambda row: row["total"],
}
VENDOR_SORT_DEFAULTS = {"vendor": "asc", "invoices": "desc", "amount": "desc"}
CAMPAIGN_SORT_FIELDS = {
    "campaign": "campaign__name",
    "year": "campaign__year",
    "team": "campaign__team__name",
    "invoices": "invoice_count",
    "amount": "total",
}
CAMPAIGN_SORT_DEFAULTS = {
    "campaign": "asc",
    "year": "desc",
    "team": "asc",
    "invoices": "desc",
    "amount": "desc",
}
TEAM_SORT_KEYS = {
    "team": lambda row: row["team"].name.lower(),
    "invoices": lambda row: row["invoice_count"],
    "amount": lambda row: row["total_spend"],
}
TEAM_SORT_DEFAULTS = {"team": "asc", "invoices": "desc", "amount": "desc"}
BUDGET_SORT_FIELDS = {
    "year": "year",
    "month": "month",
    "team": "team__name",
    "campaign": "campaign__name",
    "category": "category",
    "amount": "planned_amount",
}
BUDGET_SORT_DEFAULTS = {
    "year": "desc",
    "month": "asc",
    "team": "asc",
    "campaign": "asc",
    "category": "asc",
    "amount": "desc",
}
BUDGET_PIVOT_SORT_KEYS = {
    "team": lambda row: row["team"].lower(),
    "category": lambda row: row["category"].lower(),
    "total": lambda row: row["total"],
}
CONTRACT_SORT_FIELDS = {
    "title": "title",
    "vendor": "vendor__name",
    "team": "team__name",
    "stage": "stage",
    "end": "end_date",
    "days": "stage_changed_at",
}
CONTRACT_SORT_DEFAULTS = {
    "title": "asc",
    "vendor": "asc",
    "team": "asc",
    "stage": "asc",
    "end": "asc",
    "days": "desc",
}


def export_invoices_excel(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    queryset, _filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoices"
    sheet.append(
        [
            "Invoice number",
            "Vendor",
            "Team",
            "Campaign",
            "Category",
            "Cost bucket",
            "Invoice date",
            "Amount",
            "Currency",
            "Payment stage",
            "Days in stage",
        ]
    )
    for invoice in queryset:
        sheet.append(
            [
                invoice.invoice_number,
                invoice.vendor.name,
                invoice.team.name if invoice.team else "",
                invoice.campaign.name if invoice.campaign else "",
                invoice.category,
                invoice.cost_bucket,
                invoice.invoice_date.isoformat(),
                invoice.amount,
                invoice.currency,
                invoice.payment_stage,
                invoice.days_in_current_stage,
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="marketing-invoices.xlsx"'
    workbook.save(response)
    return response


def export_vendors_excel(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    queryset, _filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    vendor_rows = vendor_grouped_rows(queryset)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Vendors"
    sheet.append(["Vendor", "Invoice count", "Invoice numbers", "Payment stages", "Total spend"])
    for row in vendor_rows:
        sheet.append(
            [
                row["vendor"].name,
                row["invoice_count"],
                ", ".join(row["invoice_numbers"]),
                ", ".join(row["stages"]),
                row["total"],
            ]
        )
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="marketing-vendors.xlsx"'
    workbook.save(response)
    return response


def export_campaigns_excel(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    queryset, _filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    campaign_rows = (
        queryset.filter(campaign__isnull=False)
        .values("campaign__name", "campaign__year", "campaign__team__name")
        .annotate(total=Sum("amount"), invoice_count=Count("id"))
        .order_by("-total")
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Campaigns"
    sheet.append(["Campaign", "Year", "Team", "Invoice count", "Total spend"])
    for row in campaign_rows:
        sheet.append(
            [
                row["campaign__name"],
                row["campaign__year"],
                row["campaign__team__name"] or "",
                row["invoice_count"],
                row["total"],
            ]
        )
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="marketing-campaigns.xlsx"'
    workbook.save(response)
    return response


def export_workbook_excel(request):
    """Export an Excel file shaped like the source workbook (same sheet names and layout).

    This recreates the familiar workbook structure from the database — the invoice sheet,
    the wide monthly Budget projection/Actual sheet, the Market Live Spending summary, and
    the Data lookup lists — scoped to the data the current user is allowed to see.
    """
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    invoices, filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    budget_lines = filter_budget_lines_for_user(BudgetLine.objects.all(), request.user)
    if filters["year"].isdigit():
        budget_lines = budget_lines.filter(year=int(filters["year"]))
    if filters["team"].isdigit():
        budget_lines = budget_lines.filter(team_id=int(filters["team"]))

    workbook = build_workbook_style_export(invoices, budget_lines)
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="marketing-workbook.xlsx"'
    workbook.save(response)
    return response


def dashboard_report_pdf(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    queryset, filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    vendor_rows = list(
        queryset.values("vendor__name").annotate(total=Sum("amount"), invoice_count=Count("id")).order_by("-total")[:15]
    )
    stage_rows = list(queryset.values("payment_stage").annotate(invoice_count=Count("id")).order_by("payment_stage"))
    pdf_bytes = build_dashboard_summary_pdf(
        title=translate("Marketing spend summary", get_ui_lang(request)),
        generated_at=timezone.now(),
        total_spend=decimal_sum(queryset),
        invoice_count=queryset.count(),
        vendor_rows=vendor_rows,
        stage_rows=stage_rows,
        filters=filters,
        locale=pdf_locale_for_request(request),
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="marketing-dashboard-summary.pdf"'
    return response


def export_vendors_pdf(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    queryset, filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    vendor_rows = vendor_grouped_rows(queryset)
    pdf_bytes = build_vendor_report_pdf(
        title=translate("Vendor spend report", get_ui_lang(request)),
        generated_at=timezone.now(),
        vendor_rows=vendor_rows,
        filters=filters,
        locale=pdf_locale_for_request(request),
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="marketing-vendors.pdf"'
    return response


def export_campaigns_pdf(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    queryset, filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    campaign_rows = list(
        queryset.filter(campaign__isnull=False)
        .values("campaign__name", "campaign__year", "campaign__team__name")
        .annotate(total=Sum("amount"), invoice_count=Count("id"))
        .order_by("-total")
    )
    pdf_bytes = build_campaign_report_pdf(
        title=translate("Campaign spend report", get_ui_lang(request)),
        generated_at=timezone.now(),
        campaign_rows=campaign_rows,
        filters=filters,
        locale=pdf_locale_for_request(request),
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="marketing-campaigns.pdf"'
    return response


def export_contracts_excel(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    queryset, _filters = filter_contract_queryset(request, visible_contract_queryset(request))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contracts"
    sheet.append(
        [
            "Title",
            "Contract number",
            "Vendor",
            "Team",
            "Stage",
            "Start date",
            "End date",
            "Days until expiry",
            "Contract value",
            "Currency",
            "Counterparty contact",
        ]
    )
    for contract in queryset:
        sheet.append(
            [
                contract.title,
                contract.contract_number,
                contract.vendor.name,
                contract.team.name if contract.team else "",
                contract.get_stage_display(),
                contract.start_date.isoformat() if contract.start_date else "",
                contract.end_date.isoformat() if contract.end_date else "",
                contract.days_until_expiry if contract.days_until_expiry is not None else "",
                contract.amount if contract.amount is not None else "",
                contract.currency,
                contract.counterparty_contact,
            ]
        )
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="marketing-contracts.xlsx"'
    workbook.save(response)
    return response


def export_contracts_pdf(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    queryset, filters = filter_contract_queryset(request, visible_contract_queryset(request))
    pdf_bytes = build_contract_report_pdf(
        title=translate("Vendor contract report", get_ui_lang(request)),
        generated_at=timezone.now(),
        contracts=queryset,
        filters=filters,
        locale=pdf_locale_for_request(request),
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="marketing-contracts.pdf"'
    return response


def invoice_report_print(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to view reports.")
    queryset, filters = filter_invoice_queryset(request, visible_invoice_queryset(request))
    vendor_rows = list(
        queryset.values("vendor__name").annotate(total=Sum("amount"), invoice_count=Count("id")).order_by("-total")[:30]
    )
    context = {
        "generated_at": timezone.now(),
        "filters": filters,
        "total_spend": decimal_sum(queryset),
        "invoice_count": queryset.count(),
        "vendor_rows": vendor_rows,
        "stage_rows": queryset.values("payment_stage").annotate(invoice_count=Count("id")).order_by("payment_stage"),
    }
    return render(request, "marketing/print/invoice_report.html", context)


_PDF_REPORT_ROUTES = {
    "dashboard": "marketing:dashboard_report_pdf",
    "vendors": "marketing:export_vendors_pdf",
    "campaigns": "marketing:export_campaigns_pdf",
    "invoices": "marketing:invoice_report_print",
    "contracts": "marketing:export_contracts_pdf",
}


def pdf_export_wizard(request):
    if not can_export(request.user):
        return forbidden("You do not have permission to export.")
    scope_invoices = visible_invoice_queryset(request)
    if request.method == "POST":
        report = request.POST.get("report_type", "dashboard")
        params = {}
        for key in ("year", "month", "team", "business_section", "vendor", "stage", "q", "bucket"):
            value = request.POST.get(key, "").strip()
            if value:
                params[key] = value
        route = _PDF_REPORT_ROUTES.get(report, "marketing:dashboard_report_pdf")
        target = reverse(route)
        if params:
            target = f"{target}?{urlencode(params)}"
        return redirect(target)

    return render(
        request,
        "marketing/exports/pdf_wizard.html",
        {
            "report_types": [
                ("dashboard", "Dashboard summary"),
                ("vendors", "Vendor report"),
                ("campaigns", "Campaign report"),
                ("invoices", "Invoice list"),
                ("contracts", "Contract report"),
            ],
            "selected_report": request.GET.get("report", "dashboard"),
            "years": distinct_jalali_years(scope_invoices),
            "months": get_months(request),
            "teams": visible_team_queryset(request),
            "business_sections": distinct_business_sections(scope_invoices),
            "vendors": Vendor.objects.order_by("name"),
            "filters": {key: request.GET.get(key, "") for key in ("year", "month", "team", "business_section", "vendor")},
        },
    )
