from __future__ import annotations

from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from marketing.importers.excel import discover_workbook, import_marketing_workbook, load_mapping


class Command(BaseCommand):
    help = "Import marketing invoices and budget lines from the discovered XLSX workbook."

    def add_arguments(self, parser):
        parser.add_argument("--file", dest="file", help="Path to the workbook. Auto-detected if omitted.")
        parser.add_argument(
            "--mapping",
            default="docs/discovery/column_mapping.yml",
            help="Path to discovery column mapping YAML.",
        )
        parser.add_argument("--dry-run", action="store_true", help="Parse and report without writing to the database.")
        parser.add_argument(
            "--max-skipped",
            type=int,
            default=30,
            help="Maximum skipped-row reasons to print.",
        )

    def handle(self, *args, **options):
        try:
            workbook_path = discover_workbook(options["file"], base_dir=Path.cwd())
            mapping = load_mapping(options["mapping"])
        except (FileNotFoundError, ValueError) as exc:
            raise CommandError(str(exc)) from exc

        self._print_inventory(workbook_path, mapping)

        try:
            result = import_marketing_workbook(
                workbook_path,
                mapping_path=options["mapping"],
                dry_run=options["dry_run"],
            )
        except Exception as exc:
            raise CommandError(str(exc)) from exc

        prefix = "Dry run complete" if result.dry_run else "Import complete"
        self.stdout.write(self.style.SUCCESS(prefix))
        self._print_counter("Teams", result.teams)
        self._print_counter("Vendors", result.vendors)
        self._print_counter("Campaigns", result.campaigns)
        self._print_counter("Invoices", result.invoices)
        self._print_counter("Budget lines", result.budget_lines)

        if result.skipped_rows:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING(f"Skipped row reasons ({len(result.skipped_rows)} total):"))
            for skipped in result.skipped_rows[: options["max_skipped"]]:
                self.stdout.write(f"- {skipped.sheet} row {skipped.row_number}: {skipped.reason}")
            remaining = len(result.skipped_rows) - options["max_skipped"]
            if remaining > 0:
                self.stdout.write(f"- ... {remaining} more skipped reasons not shown")

    def _print_inventory(self, workbook_path: Path, mapping: dict):
        self.stdout.write(f"Workbook: {workbook_path}")
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            self.stdout.write(f"Detected sheets: {', '.join(workbook.sheetnames)}")
        finally:
            workbook.close()

        invoices = mapping.get("sheets", {}).get("invoices", {})
        budget = mapping.get("sheets", {}).get("budget", {})
        self.stdout.write(
            "Invoice mapping: "
            f"{invoices.get('actual_sheet_name')} header row {invoices.get('header_row')}, "
            f"rows {invoices.get('data_start_row')}-{invoices.get('data_end_row')}"
        )
        self.stdout.write(
            "Budget mapping: "
            f"{budget.get('actual_sheet_name')} header row {budget.get('header_row')}, "
            f"rows {budget.get('data_start_row')}-{budget.get('data_end_row')}"
        )

    def _print_counter(self, label, counter):
        self.stdout.write(f"{label}: created={counter.created}, updated={counter.updated}, skipped={counter.skipped}")
