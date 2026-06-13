from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import yaml
from django.conf import settings
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

from marketing.cost_buckets import (
    detect_cost_bucket_from_text,
    infer_cost_bucket_from_pseudo_team_name,
    is_pseudo_team_name,
    parent_team_name_for_bucket,
)
from marketing.jalali import normalize_digits, parse_jalali_date_text
from marketing.models import (
    BudgetLine,
    Campaign,
    CostBucket,
    Invoice,
    PaymentStage,
    Team,
    TeamAlias,
    Vendor,
    normalize_name,
)


class DryRunRollback(Exception):
    pass


@dataclass
class ImportCounter:
    created: int = 0
    updated: int = 0
    skipped: int = 0


@dataclass
class SkippedRow:
    sheet: str
    row_number: int
    reason: str


@dataclass
class ImportResult:
    workbook: Path
    dry_run: bool
    teams: ImportCounter = field(default_factory=ImportCounter)
    vendors: ImportCounter = field(default_factory=ImportCounter)
    campaigns: ImportCounter = field(default_factory=ImportCounter)
    invoices: ImportCounter = field(default_factory=ImportCounter)
    budget_lines: ImportCounter = field(default_factory=ImportCounter)
    skipped_rows: list[SkippedRow] = field(default_factory=list)
    dry_run_teams: dict[str, Team] = field(default_factory=dict, repr=False)
    dry_run_vendors: dict[str, Vendor] = field(default_factory=dict, repr=False)
    dry_run_campaigns: dict[tuple[str, int, str | None], Campaign] = field(default_factory=dict, repr=False)

    def skip(self, sheet: str, row_number: int, reason: str, counter: ImportCounter | None = None) -> None:
        if counter is not None:
            counter.skipped += 1
        self.skipped_rows.append(SkippedRow(sheet=sheet, row_number=row_number, reason=reason))


def discover_workbook(explicit_path: str | Path | None = None, base_dir: Path | None = None) -> Path:
    if explicit_path:
        path = Path(explicit_path)
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")
        return path

    base = base_dir or Path.cwd()
    candidates = [
        *base.glob("*.xlsx"),
        *base.glob("data/*.xlsx"),
        *base.glob("imports/*.xlsx"),
    ]
    candidates = [path for path in candidates if not path.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError("No .xlsx workbook found in project root, data/, or imports/.")
    if len(candidates) > 1:
        formatted = "\n".join(f"- {path}" for path in candidates)
        raise ValueError(f"Multiple .xlsx workbooks found. Pass --file explicitly:\n{formatted}")
    return candidates[0]


def load_mapping(mapping_path: str | Path = "docs/discovery/column_mapping.yml") -> dict[str, Any]:
    path = Path(mapping_path)
    if not path.exists():
        raise FileNotFoundError(
            f"Mapping file not found: {path}. Run discovery first or provide --mapping with a valid mapping YAML."
        )
    with path.open(encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value).strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    cleaned = re.sub(r"[,\s]", "", str(value))
    cleaned = re.sub(r"(IRR|Rial|ریال|تومان)", "", cleaned, flags=re.IGNORECASE)
    if cleaned == "":
        return None
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def parse_year(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except ValueError:
        return None


def parse_excel_date(value: Any, workbook_epoch) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, int | float):
        try:
            parsed = from_excel(value, workbook_epoch)
        except (TypeError, ValueError):
            return None
        return parsed.date() if isinstance(parsed, datetime) else parsed
    text = normalize_digits(cell_to_text(value))
    # Values like 1405/01/10 are Jalali/Shamsi. Parse those before trying
    # Gregorian formats; otherwise Python would accept year 1405 as Gregorian.
    jalali_date = parse_jalali_date_text(text)
    if jalali_date:
        return jalali_date
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def aware_datetime_from_date(value: date | None) -> datetime | None:
    """Convert a date into a timezone-aware datetime (anchored at midday to avoid TZ edge cases)."""
    if value is None:
        return None
    naive = datetime(value.year, value.month, value.day, 12, 0)
    if timezone.is_naive(naive):
        return timezone.make_aware(naive, timezone.get_current_timezone())
    return naive


def normalize_currency(value: Any) -> str:
    text = cell_to_text(value)
    if not text:
        return settings.DEFAULT_CURRENCY
    normalized = normalize_name(text)
    if normalized in {"rial", "ریال", "irr"}:
        return "IRR"
    return text[:16]


def header_map_for_sheet(sheet, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    seen: dict[str, int] = {}
    for index, cell in enumerate(sheet[header_row], start=1):
        value = cell_to_text(cell.value)
        if not value:
            value = f"blank_{get_column_letter(index)}"
        seen[value] = seen.get(value, 0) + 1
        key = value if seen[value] == 1 else f"{value} ({seen[value]})"
        headers[key] = index
    return headers


def row_data(sheet, row_number: int, headers: dict[str, int]) -> dict[str, Any]:
    data = {}
    for header, column_index in headers.items():
        data[header] = json_safe(sheet.cell(row=row_number, column=column_index).value)
    return data


def mapped_value(row: dict[str, Any], columns: dict[str, str | None], key: str) -> Any:
    column = columns.get(key)
    if not column:
        return None
    return row.get(column)


# Canonical team names for workbook variants that mean the same team. Keyed by the normalized
# (casefolded, whitespace-collapsed) raw name so they resolve at import time independent of DB
# state. The DB-backed TeamAlias table remains available for admin-defined aliases.
TEAM_NAME_ALIASES: dict[str, str] = {
    normalize_name("Operation & Analysis"): "Ops & Analytics",
    normalize_name("Brand (PR & Social & CSR)"): "Brand",
}


def canonical_team_name(name: str) -> str:
    """Resolve known workbook spelling variants to a single canonical team name."""
    text = " ".join((name or "").split())
    return TEAM_NAME_ALIASES.get(normalize_name(text), text)


def get_or_create_team(name: str, result: ImportResult, dry_run: bool) -> Team:
    if is_pseudo_team_name(name):
        raise ValueError(f"Refusing to create pseudo-team from cost-bucket label: {name!r}")
    name = canonical_team_name(name)
    alias = TeamAlias.objects.select_related("team").filter(
        normalized_raw_name=normalize_name(name),
        is_active=True,
        team__is_active=True,
    ).first()
    if alias:
        return alias.team

    slug = slugify(name, allow_unicode=True) or normalize_name(name).replace(" ", "-")
    team = Team.objects.filter(slug=slug).first() or Team.objects.filter(name=name).first()
    if team:
        return team
    if dry_run:
        if slug in result.dry_run_teams:
            return result.dry_run_teams[slug]
        result.teams.created += 1
        team = Team(name=name, slug=slug)
        result.dry_run_teams[slug] = team
        return team
    team = Team.objects.create(name=name, slug=slug)
    result.teams.created += 1
    return team


def resolve_import_team(
    team_name: str,
    cost_bucket: str,
    result: ImportResult,
    dry_run: bool,
) -> Team | None:
    parent_name = parent_team_name_for_bucket(cost_bucket)
    if parent_name:
        return get_or_create_team(parent_name, result, dry_run)
    if team_name and not is_pseudo_team_name(team_name):
        return get_or_create_team(team_name, result, dry_run)
    return None


def resolve_budget_team(
    team_name: str,
    category: str,
    description: str,
    result: ImportResult,
    dry_run: bool,
) -> Team | None:
    context_text = " ".join(part for part in (team_name, category, description) if part)
    bucket = detect_cost_bucket_from_text(context_text)
    if bucket:
        parent_name = parent_team_name_for_bucket(bucket)
        if parent_name:
            return get_or_create_team(parent_name, result, dry_run)
    if team_name and not is_pseudo_team_name(team_name):
        return get_or_create_team(team_name, result, dry_run)
    return None


def merge_aliased_teams() -> None:
    """Fold duplicate workbook team variants into their canonical team.

    Robust to import order: runs every load so a fresh DB (where the seed migration found no
    teams) still ends up with a single canonical team and a recorded TeamAlias row.
    """
    for raw_normalized, canonical_name in TEAM_NAME_ALIASES.items():
        canonical = Team.objects.filter(name=canonical_name).first()
        if canonical is None:
            continue
        duplicates = [
            team
            for team in Team.objects.exclude(pk=canonical.pk)
            if normalize_name(team.name) == raw_normalized
        ]
        for alias_team in duplicates:
            TeamAlias.objects.update_or_create(
                raw_name=alias_team.name,
                defaults={
                    "normalized_raw_name": normalize_name(alias_team.name),
                    "team": canonical,
                    "is_active": True,
                    "notes": "Auto-merged workbook spelling variant during import.",
                },
            )
            for campaign in Campaign.objects.filter(team=alias_team):
                existing = (
                    Campaign.objects.filter(name=campaign.name, year=campaign.year, team=canonical)
                    .exclude(pk=campaign.pk)
                    .first()
                )
                if existing:
                    Invoice.objects.filter(campaign=campaign).update(campaign=existing)
                    BudgetLine.objects.filter(campaign=campaign).update(campaign=existing)
                    campaign.delete()
                else:
                    campaign.team = canonical
                    campaign.save(update_fields=["team"])
            Invoice.objects.filter(team=alias_team).update(team=canonical)
            BudgetLine.objects.filter(team=alias_team).update(team=canonical)
            alias_team.is_active = False
            alias_team.save(update_fields=["is_active"])


def deactivate_reassigned_pseudo_teams() -> None:
    """Hide legacy pseudo-teams and move any remaining rows onto parent teams."""
    for team in Team.objects.filter(is_active=True):
        if not is_pseudo_team_name(team.name):
            continue
        bucket = infer_cost_bucket_from_pseudo_team_name(team.name)
        parent_name = parent_team_name_for_bucket(bucket) if bucket else None
        if parent_name and bucket:
            parent = Team.objects.filter(name=parent_name).first()
            if parent:
                Invoice.objects.filter(team=team).update(team=parent, cost_bucket=bucket)
                BudgetLine.objects.filter(team=team).update(team=parent)
        team.is_active = False
        team.save(update_fields=["is_active"])


def get_or_create_vendor(name: str, result: ImportResult, dry_run: bool) -> Vendor:
    normalized = normalize_name(name)
    vendor = Vendor.objects.filter(normalized_name=normalized).first()
    if vendor:
        return vendor
    if dry_run:
        if normalized in result.dry_run_vendors:
            return result.dry_run_vendors[normalized]
        result.vendors.created += 1
        vendor = Vendor(name=name, normalized_name=normalized)
        result.dry_run_vendors[normalized] = vendor
        return vendor
    vendor = Vendor.objects.create(name=name, normalized_name=normalized)
    result.vendors.created += 1
    return vendor


# Canonical spellings for free-text campaign names coming from the workbook, keyed by the
# normalized (casefolded, whitespace-collapsed) form so variants like "on going" / "on-going" /
# "ongoing" all resolve to a single consistent display value.
CAMPAIGN_NAME_ALIASES = {
    "on going": "Ongoing",
    "ongoing": "Ongoing",
    "on going campaign": "Ongoing",
}


def canonical_campaign_name(name: str) -> str:
    """Return a consistent spelling for known free-text campaign-name variants."""
    text = " ".join((name or "").split())
    if not text:
        return ""
    return CAMPAIGN_NAME_ALIASES.get(normalize_name(text), text)


def get_or_create_campaign(
    name: str,
    year: int,
    team: Team | None,
    result: ImportResult,
    dry_run: bool,
) -> Campaign | None:
    name = canonical_campaign_name(name)
    if not name:
        return None
    campaign = Campaign.objects.filter(name=name, year=year, team=team if team and team.pk else None).first()
    if campaign:
        return campaign
    if dry_run:
        team_key = team.slug if team else None
        cache_key = (name, year, team_key)
        if cache_key in result.dry_run_campaigns:
            return result.dry_run_campaigns[cache_key]
        result.campaigns.created += 1
        campaign = Campaign(name=name, year=year, team=team if team and team.pk else None)
        result.dry_run_campaigns[cache_key] = campaign
        return campaign
    campaign = Campaign.objects.create(name=name, year=year, team=team if team and team.pk else None)
    result.campaigns.created += 1
    return campaign


def detect_cost_bucket(row: dict[str, Any], bucket_mapping: dict[str, Any]) -> str:
    default = bucket_mapping.get("default", CostBucket.TEAM)
    for rule in bucket_mapping.get("rules", []):
        keywords = [normalize_name(keyword) for keyword in rule.get("when_any_source_column_contains", [])]
        source_columns = rule.get("source_columns", [])
        text = " ".join(cell_to_text(row.get(column)) for column in source_columns)
        normalized_text = normalize_name(text)
        if any(keyword and keyword in normalized_text for keyword in keywords):
            return rule.get("bucket", default)
    return default


def map_payment_stage(value: Any, stage_mapping: dict[str, str]) -> str:
    text = cell_to_text(value)
    if not text:
        return PaymentStage.DRAFT
    if text in stage_mapping:
        return stage_mapping[text]
    normalized = normalize_name(text)
    for source, target in stage_mapping.items():
        if normalize_name(source) == normalized:
            return target
    if normalized in {"paid", "پرداخت شده"}:
        return PaymentStage.PAID
    if normalized in {"finance", "مالی"}:
        return PaymentStage.FINANCE_REVIEW
    return PaymentStage.SUBMITTED


def duplicate_invoice_vendor_keys(
    sheet,
    invoice_mapping: dict[str, Any],
    headers: dict[str, int],
) -> set[tuple[str, str]]:
    columns = invoice_mapping["columns"]
    counts: dict[tuple[str, str], int] = {}
    for row_number in range(invoice_mapping["data_start_row"], invoice_mapping["data_end_row"] + 1):
        raw = row_data(sheet, row_number, headers)
        invoice_number = cell_to_text(mapped_value(raw, columns, "invoice_number"))
        vendor_name = cell_to_text(mapped_value(raw, columns, "vendor_name"))
        if not invoice_number or not vendor_name:
            continue
        key = (invoice_number, normalize_name(vendor_name))
        counts[key] = counts.get(key, 0) + 1
    return {key for key, count in counts.items() if count > 1}


def find_existing_invoice(
    vendor: Vendor,
    invoice_number: str,
    source_sheet: str,
    source_row: int,
    *,
    allow_invoice_vendor_fallback: bool,
) -> Invoice | None:
    if vendor.pk:
        source_match = Invoice.objects.filter(
            source_sheet=source_sheet,
            source_row_number=source_row,
            vendor=vendor,
            invoice_number=invoice_number,
        ).first()
        if source_match:
            return source_match

        if allow_invoice_vendor_fallback:
            candidates = Invoice.objects.filter(vendor=vendor, invoice_number=invoice_number)
            if candidates.count() == 1:
                return candidates.first()
    return None


def update_or_create_invoice(
    *,
    vendor: Vendor,
    team: Team | None,
    campaign: Campaign | None,
    invoice_number: str,
    category: str,
    cost_bucket: str,
    description: str,
    invoice_date: date,
    amount: Decimal,
    currency: str,
    payment_stage: str,
    source_sheet: str,
    source_row: int,
    raw_data: dict[str, Any],
    result: ImportResult,
    dry_run: bool,
    stage_changed_at: datetime | None = None,
    allow_invoice_vendor_fallback: bool = True,
) -> None:
    existing = find_existing_invoice(
        vendor,
        invoice_number,
        source_sheet,
        source_row,
        allow_invoice_vendor_fallback=allow_invoice_vendor_fallback,
    )
    if dry_run:
        if existing:
            result.invoices.updated += 1
        else:
            result.invoices.created += 1
        return

    values = {
        "team": team,
        "campaign": campaign,
        "category": category,
        "cost_bucket": cost_bucket,
        "description": description,
        "invoice_date": invoice_date,
        "amount": amount,
        "currency": currency,
        "payment_stage": payment_stage,
        "source_sheet": source_sheet,
        "source_row_number": source_row,
        "raw_data_json": raw_data,
    }
    # Aging ("days in current stage") must reflect when the invoice actually entered its stage.
    # The workbook records when MKT sent the invoice to Finance, so use that as the stage start
    # for finance-review rows; otherwise Invoice.save() would stamp it with the import time.
    if stage_changed_at is not None and payment_stage == PaymentStage.FINANCE_REVIEW:
        values["stage_changed_at"] = stage_changed_at
    if existing:
        for key, value in values.items():
            setattr(existing, key, value)
        existing.save()
        result.invoices.updated += 1
        return

    Invoice.objects.create(invoice_number=invoice_number, vendor=vendor, **values)
    result.invoices.created += 1


def import_invoices(workbook, mapping: dict[str, Any], result: ImportResult, dry_run: bool) -> None:
    invoice_mapping = mapping["sheets"]["invoices"]
    sheet_name = invoice_mapping["actual_sheet_name"]
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Invoice sheet not found: {sheet_name}")

    sheet = workbook[sheet_name]
    headers = header_map_for_sheet(sheet, invoice_mapping["header_row"])
    columns = invoice_mapping["columns"]
    bucket_mapping = invoice_mapping.get("derived_values", {}).get("cost_bucket", {})
    stage_mapping = invoice_mapping.get("derived_values", {}).get("payment_stage_mapping", {})
    currency = invoice_mapping.get("derived_values", {}).get("currency", settings.DEFAULT_CURRENCY)
    duplicate_invoice_keys = duplicate_invoice_vendor_keys(sheet, invoice_mapping, headers)

    for row_number in range(invoice_mapping["data_start_row"], invoice_mapping["data_end_row"] + 1):
        raw = row_data(sheet, row_number, headers)
        if not any(cell_to_text(value) for value in raw.values()):
            result.skip(sheet_name, row_number, "empty row", result.invoices)
            continue

        invoice_number = cell_to_text(mapped_value(raw, columns, "invoice_number"))
        vendor_name = cell_to_text(mapped_value(raw, columns, "vendor_name"))
        amount = parse_decimal(mapped_value(raw, columns, "amount"))
        invoice_date = parse_excel_date(
            mapped_value(raw, columns, "jalali_invoice_date_text_candidate"),
            workbook.epoch,
        )
        invoice_date = invoice_date or parse_excel_date(
            mapped_value(raw, columns, "invoice_date_gregorian_serial"),
            workbook.epoch,
        )
        invoice_date = invoice_date or parse_excel_date(
            mapped_value(raw, columns, "invoice_date_jalali_serial"),
            workbook.epoch,
        )
        year = parse_year(mapped_value(raw, columns, "year")) or (invoice_date.year if invoice_date else None)
        category = cell_to_text(mapped_value(raw, columns, "category")) or "Uncategorized"

        missing = []
        if not invoice_number:
            missing.append("invoice_number")
        if not vendor_name:
            missing.append("vendor_name")
        if amount is None:
            missing.append("amount")
        if invoice_date is None:
            missing.append("invoice_date")
        if year is None:
            missing.append("year")
        if missing:
            result.skip(sheet_name, row_number, f"missing required fields: {', '.join(missing)}", result.invoices)
            continue

        vendor = get_or_create_vendor(vendor_name, result, dry_run)
        invoice_key = (invoice_number, normalize_name(vendor_name))
        team_name = cell_to_text(mapped_value(raw, columns, "team"))
        cost_bucket = detect_cost_bucket(raw, bucket_mapping)
        inferred_bucket = infer_cost_bucket_from_pseudo_team_name(team_name)
        if inferred_bucket:
            cost_bucket = inferred_bucket
        team = resolve_import_team(team_name, cost_bucket, result, dry_run)
        campaign_name = cell_to_text(mapped_value(raw, columns, "campaign_name"))
        campaign = get_or_create_campaign(campaign_name, year, team, result, dry_run) if year else None
        description = cell_to_text(mapped_value(raw, columns, "description"))
        payment_stage = map_payment_stage(mapped_value(raw, columns, "payment_stage"), stage_mapping)
        finance_sent_date = parse_excel_date(
            mapped_value(raw, columns, "finance_sent_date_gregorian_serial"),
            workbook.epoch,
        ) or parse_excel_date(
            mapped_value(raw, columns, "finance_sent_date_jalali_serial"),
            workbook.epoch,
        )

        update_or_create_invoice(
            vendor=vendor,
            team=team,
            campaign=campaign,
            invoice_number=invoice_number,
            category=category,
            cost_bucket=cost_bucket,
            description=description,
            invoice_date=invoice_date,
            amount=amount,
            currency=currency,
            payment_stage=payment_stage,
            source_sheet=sheet_name,
            source_row=row_number,
            raw_data=raw,
            result=result,
            dry_run=dry_run,
            stage_changed_at=aware_datetime_from_date(finance_sent_date),
            allow_invoice_vendor_fallback=invoice_key not in duplicate_invoice_keys,
        )


def find_existing_budget_line(
    *,
    source_sheet: str,
    source_row: int,
    year: int,
    month: int,
) -> BudgetLine | None:
    # Idempotency key is the workbook source row + month only. Team/category are derived from the
    # row and can change between imports (e.g. team alias merges), so they must NOT be part of the
    # lookup — otherwise a re-import would create a duplicate instead of updating the existing line.
    if source_row is None:
        return None
    return BudgetLine.objects.filter(
        source_sheet=source_sheet,
        source_row_number=source_row,
        year=year,
        month=month,
    ).first()


def update_or_create_budget_line(
    *,
    year: int,
    month: int,
    team: Team | None,
    category: str,
    planned_amount: Decimal,
    currency: str,
    source_sheet: str,
    source_row: int,
    raw_data: dict[str, Any],
    result: ImportResult,
    dry_run: bool,
) -> None:
    existing = find_existing_budget_line(
        source_sheet=source_sheet,
        source_row=source_row,
        year=year,
        month=month,
    )
    if dry_run:
        if existing:
            result.budget_lines.updated += 1
        else:
            result.budget_lines.created += 1
        return

    values = {
        "team": team,
        "category": category,
        "planned_amount": planned_amount,
        "currency": currency,
        "source_sheet": source_sheet,
        "source_row_number": source_row,
        "raw_data_json": raw_data,
    }
    if existing:
        for key, value in values.items():
            setattr(existing, key, value)
        existing.save()
        result.budget_lines.updated += 1
        return

    BudgetLine.objects.create(year=year, month=month, **values)
    result.budget_lines.created += 1


def import_budget_lines(workbook, mapping: dict[str, Any], result: ImportResult, dry_run: bool) -> None:
    budget_mapping = mapping["sheets"]["budget"]
    sheet_name = budget_mapping["actual_sheet_name"]
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Budget sheet not found: {sheet_name}")

    sheet = workbook[sheet_name]
    header_row = budget_mapping["header_row"]
    headers = header_map_for_sheet(sheet, header_row)
    context_columns = budget_mapping["row_context_columns"]
    year = int(budget_mapping["budget_line_mapping"]["year"])

    for row_number in range(budget_mapping["data_start_row"], budget_mapping["data_end_row"] + 1):
        raw = row_data(sheet, row_number, headers)
        if not any(cell_to_text(value) for value in raw.values()):
            result.skip(sheet_name, row_number, "empty row", result.budget_lines)
            continue

        team_name = cell_to_text(raw.get(context_columns["team"]))
        title = cell_to_text(raw.get(context_columns["category_or_title"]))
        description = cell_to_text(raw.get(context_columns["description"]))
        category = title or description
        if not team_name or not category:
            result.skip(sheet_name, row_number, "missing team or category/title", result.budget_lines)
            continue

        team = resolve_budget_team(team_name, category, description, result, dry_run)
        if team is None:
            result.skip(sheet_name, row_number, "could not resolve team (pseudo-team label only)", result.budget_lines)
            continue
        currency = normalize_currency(raw.get(context_columns["currency"]))

        for month_info in budget_mapping["monthly_columns"]:
            month = int(month_info["month_number"])
            projection_cell = sheet[f"{month_info['projection_column']}{row_number}"]
            actual_cell = sheet[f"{month_info['actual_column']}{row_number}"]
            planned_amount = parse_decimal(projection_cell.value)
            if planned_amount is None:
                result.skip(
                    sheet_name,
                    row_number,
                    f"missing planned amount for month {month}",
                    result.budget_lines,
                )
                continue

            monthly_raw = {
                **raw,
                "_import_month": month,
                "_import_month_label": month_info["month_label"],
                "_projection_column": month_info["projection_column"],
                "_actual_column": month_info["actual_column"],
                "_actual_amount_reference": json_safe(actual_cell.value),
            }
            update_or_create_budget_line(
                year=year,
                month=month,
                team=team,
                category=category,
                planned_amount=planned_amount,
                currency=currency,
                source_sheet=sheet_name,
                source_row=row_number,
                raw_data=monthly_raw,
                result=result,
                dry_run=dry_run,
            )


def import_marketing_workbook(
    workbook_path: str | Path,
    *,
    mapping_path: str | Path = "docs/discovery/column_mapping.yml",
    dry_run: bool = False,
) -> ImportResult:
    workbook_file = Path(workbook_path)
    mapping = load_mapping(mapping_path)
    result = ImportResult(workbook=workbook_file, dry_run=dry_run)

    workbook = load_workbook(workbook_file, data_only=True)
    try:
        with transaction.atomic():
            import_invoices(workbook, mapping, result, dry_run)
            import_budget_lines(workbook, mapping, result, dry_run)
            if dry_run:
                raise DryRunRollback
    except DryRunRollback:
        pass
    finally:
        workbook.close()
    if not dry_run:
        merge_aliased_teams()
        deactivate_reassigned_pseudo_teams()
    return result
