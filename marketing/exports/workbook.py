from __future__ import annotations

import re
from collections import defaultdict
from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from django.db.models import QuerySet
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from marketing.jalali import JALALI_MONTHS, format_jalali_date, gregorian_to_jalali
from marketing.models import BudgetLine, Invoice, PaymentStage

ZERO = Decimal("0")
MONTHS = [(number, latin) for number, _persian, latin in JALALI_MONTHS]
MONEY_FORMAT = '#,##0'
# 1-based money column indexes per exported sheet (used during final styling pass).
_SHEET_MONEY_COLUMNS: dict[str, set[int]] = {}


def build_workbook_style_export(invoices: QuerySet[Invoice], budget_lines: QuerySet[BudgetLine]) -> Workbook:
    """Build an Excel file shaped like the source workbook, using scoped DB data.

    This is not a byte-for-byte recreation of the imported workbook. It is a clean
    database export with familiar sheet names and wide monthly budget/actual views.
    Output is validated for Microsoft Excel and Google Sheets, the primary consumers.
    """
    invoice_rows = list(invoices.select_related("vendor", "team", "campaign").order_by("invoice_date", "id"))
    budget_rows = list(
        budget_lines.select_related("team", "campaign").order_by("year", "team__name", "category", "month")
    )

    workbook = Workbook()
    _SHEET_MONEY_COLUMNS.clear()
    _add_invoice_sheet(workbook.active, invoice_rows)
    _add_budget_sheet(workbook.create_sheet("Budget"), budget_rows, invoice_rows)
    _add_market_live_spending_sheet(workbook.create_sheet("Market Live Spending"), budget_rows, invoice_rows)
    _add_data_sheet(workbook.create_sheet("Data"), budget_rows, invoice_rows)
    _finalize_workbook(workbook)
    return workbook


def _add_invoice_sheet(sheet: Worksheet, invoices: Iterable[Invoice]) -> None:
    sheet.title = "Marketing Spend Input"
    headers = [
        "Year",
        "Month",
        "Cost Center",
        "MKT Team",
        "Sub Team",
        "Requester",
        "Budget Line",
        "Campaign Name",
        "Business Section",
        "Vendor Name",
        "Description",
        "invoice date in Jalali",
        "invoice date in gregorian",
        "Invoice Number",
        "Invoice Amount (IRR)",
        "Action Cost (iRR)",
        "Tax Amount (IRR)",
        "Insurance Rate",
        "Insurance Amount (IRR)",
        "Paid (IRR)",
        "MKT Confirmation stage",
        "MKT to Finance sent date",
        "MKT to Finance sent date in gregorian",
        "payment state",
        "lead Time",
    ]
    _append_header(sheet, headers)

    for invoice in invoices:
        raw = invoice.raw_data_json or {}
        if invoice.invoice_date:
            jalali_year, jalali_month, _day = gregorian_to_jalali(
                invoice.invoice_date.year,
                invoice.invoice_date.month,
                invoice.invoice_date.day,
            )
            month_label = dict(MONTHS).get(jalali_month, str(jalali_month))
            jalali_text = format_jalali_date(invoice.invoice_date)
            gregorian_text = invoice.invoice_date.isoformat()
        else:
            jalali_year, month_label, jalali_text, gregorian_text = "", "", "", ""
        sheet.append(_row_values([
            _raw(raw, "Year", jalali_year),
            _raw(raw, "Month", month_label),
            _raw(raw, "Cost Center", ""),
            invoice.team.name if invoice.team else "",
            _raw(raw, "Sub Team", ""),
            _raw(raw, "Requester", ""),
            invoice.category,
            invoice.campaign.name if invoice.campaign else "",
            _raw(raw, "Business Section", ""),
            invoice.vendor.name if invoice.vendor_id else "",
            invoice.description,
            _raw(raw, "invoice date in Jalali", jalali_text),
            gregorian_text,
            invoice.invoice_number,
            _excel_number(invoice.amount),
            _excel_number(_raw(raw, "Action Cost (iRR)", "")),
            _excel_number(_raw(raw, "Tax Amount (IRR)", "")),
            _raw(raw, "Insurance Rate", ""),
            _excel_number(_raw(raw, "Insurance Amount (IRR)", "")),
            _excel_number(
                _raw(raw, "Paid (IRR)", invoice.amount if invoice.payment_stage == PaymentStage.PAID else "")
            ),
            _raw(raw, "MKT Confirmation stage", ""),
            _raw(raw, "MKT to Finance sent date", ""),
            _raw(raw, "MKT to Finance sent date in gregorian", ""),
            _payment_stage_for_excel(invoice),
            _raw(raw, "lead Time", invoice.days_in_current_stage),
        ]))

    _SHEET_MONEY_COLUMNS[sheet.title] = {15, 16, 17, 19, 20}


def _add_budget_sheet(sheet: Worksheet, budget_lines: Iterable[BudgetLine], invoices: Iterable[Invoice]) -> None:
    headers = [
        "Team",
        "Sub Team",
        "Team Lead",
        "Org. Vertical",
        "Title",
        "Description",
        "Unit Price",
        "Total projection",
        "Total Actual",
    ]
    for _month_number, month_label in MONTHS:
        headers.extend([f"{month_label} projection", f"{month_label} Actual"])
    _append_header(sheet, headers)

    rows = _budget_actual_rows(budget_lines, invoices)
    for row in rows:
        values = [
            row["team"],
            row["sub_team"],
            row["team_lead"],
            row["org_vertical"],
            row["category"],
            row["description"],
            row["currency"],
            _excel_number(row["planned_total"]),
            _excel_number(row["actual_total"]),
        ]
        for month_number, _month_label in MONTHS:
            values.append(_excel_number(row["planned"][month_number]))
            values.append(_excel_number(row["actual"][month_number]))
        sheet.append(_row_values(values))

    _SHEET_MONEY_COLUMNS[sheet.title] = set(range(8, len(headers) + 1))


def _add_market_live_spending_sheet(
    sheet: Worksheet,
    budget_lines: Iterable[BudgetLine],
    invoices: Iterable[Invoice],
) -> None:
    headers = ["Teams", "Title", *[month_label for _month_number, month_label in MONTHS], "Total"]
    _append_header(sheet, headers)

    planned = defaultdict(lambda: {month: ZERO for month, _label in MONTHS})
    actual = defaultdict(lambda: {month: ZERO for month, _label in MONTHS})
    labels: set[str] = set()

    for line in budget_lines:
        label = _team_label_from_budget(line)
        labels.add(label)
        if line.month:
            planned[label][line.month] += line.planned_amount or ZERO

    for invoice in invoices:
        if not invoice.invoice_date:
            continue
        label = _team_label_from_invoice(invoice)
        labels.add(label)
        actual[label][invoice.jalali_month] += invoice.amount or ZERO

    for label in sorted(labels):
        projection_values = planned[label]
        actual_values = actual[label]
        remaining_values = {
            month_number: projection_values[month_number] - actual_values[month_number]
            for month_number, _month_label in MONTHS
        }
        deviation_values = {
            month_number: actual_values[month_number] - projection_values[month_number]
            for month_number, _month_label in MONTHS
        }
        _append_month_summary_row(sheet, label, "projection", projection_values)
        _append_month_summary_row(sheet, label, "Actual", actual_values)
        _append_month_summary_row(sheet, label, "Remaining", remaining_values)
        _append_month_summary_row(sheet, label, "Deviation", deviation_values)

    _SHEET_MONEY_COLUMNS[sheet.title] = set(range(3, len(headers) + 1))


def _add_data_sheet(sheet: Worksheet, budget_lines: Iterable[BudgetLine], invoices: Iterable[Invoice]) -> None:
    headers = ["requester", "Vendor List", "Vendor list unique", "Title", "Title list unique", "Month list", "sub team"]
    _append_header(sheet, headers)

    requesters = sorted(
        {_raw(invoice.raw_data_json or {}, "Requester", "") for invoice in invoices if invoice.raw_data_json}
    )
    vendors = sorted({invoice.vendor.name for invoice in invoices if invoice.vendor_id})
    categories = sorted(
        {invoice.category for invoice in invoices if invoice.category}
        | {line.category for line in budget_lines if line.category}
    )
    sub_teams = sorted(
        {_raw(invoice.raw_data_json or {}, "Sub Team", "") for invoice in invoices if invoice.raw_data_json}
        | {_raw(line.raw_data_json or {}, "Sub Team", "") for line in budget_lines if line.raw_data_json}
    )
    month_labels = [month_label for _month_number, month_label in MONTHS]

    rows_count = max(len(requesters), len(vendors), len(categories), len(month_labels), len(sub_teams), 1)
    for index in range(rows_count):
        vendor = _list_value(vendors, index)
        category = _list_value(categories, index)
        sheet.append(_row_values([
            _list_value(requesters, index),
            vendor,
            vendor,
            category,
            category,
            _list_value(month_labels, index),
            _list_value(sub_teams, index),
        ]))

    _SHEET_MONEY_COLUMNS[sheet.title] = set()


def _budget_actual_rows(budget_lines: Iterable[BudgetLine], invoices: Iterable[Invoice]) -> list[dict]:
    rows: dict[tuple[str, str], dict] = {}

    for line in budget_lines:
        key = (_team_label_from_budget(line), line.category or "Uncategorized")
        row = rows.setdefault(key, _empty_budget_row(key[0], key[1]))
        raw = line.raw_data_json or {}
        row["sub_team"] = row["sub_team"] or _raw(raw, "Sub Team", "")
        row["team_lead"] = row["team_lead"] or _raw(raw, "Team Lead", "")
        row["org_vertical"] = row["org_vertical"] or _raw(raw, "Org. Vertical", "")
        row["description"] = row["description"] or _raw(raw, "Description", "")
        row["currency"] = line.currency or row["currency"]
        if line.month:
            row["planned"][line.month] += line.planned_amount or ZERO
        row["planned_total"] += line.planned_amount or ZERO

    for invoice in invoices:
        key = (_team_label_from_invoice(invoice), invoice.category or "Uncategorized")
        row = rows.setdefault(key, _empty_budget_row(key[0], key[1]))
        row["currency"] = invoice.currency or row["currency"]
        if invoice.invoice_date:
            row["actual"][invoice.jalali_month] += invoice.amount or ZERO
        row["actual_total"] += invoice.amount or ZERO

    return sorted(rows.values(), key=lambda item: (item["team"], item["category"]))


def _empty_budget_row(team: str, category: str) -> dict:
    return {
        "team": team,
        "sub_team": "",
        "team_lead": "",
        "org_vertical": "",
        "category": category,
        "description": "",
        "currency": "IRR",
        "planned": {month_number: ZERO for month_number, _label in MONTHS},
        "actual": {month_number: ZERO for month_number, _label in MONTHS},
        "planned_total": ZERO,
        "actual_total": ZERO,
    }


def _append_month_summary_row(sheet: Worksheet, team: str, title: str, values_by_month: dict[int, Decimal]) -> None:
    month_values = [values_by_month[month_number] for month_number, _month_label in MONTHS]
    sheet.append(_row_values([
        team,
        title,
        *[_excel_number(value) for value in month_values],
        _excel_number(sum(month_values, ZERO)),
    ]))


def _append_header(sheet: Worksheet, headers: list[str]) -> None:
    sheet.append(_row_values(headers))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center")


def _finalize_workbook(workbook: Workbook) -> None:
    """Apply filters, widths, and number formats once every sheet is populated."""
    for sheet in workbook.worksheets:
        _style_table(sheet, money_columns=_SHEET_MONEY_COLUMNS.get(sheet.title, set()))


def _style_table(sheet: Worksheet, *, money_columns: set[int] | None = None) -> None:
    money_columns = money_columns or set()
    max_row = sheet.max_row or 1
    max_column = sheet.max_column or 1

    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.value = _normalize_cell_value(cell.value)
            if cell.row == 1:
                continue
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column not in money_columns)
            if cell.column in money_columns and isinstance(cell.value, int | float):
                cell.number_format = MONEY_FORMAT

    if max_row >= 2 and max_column >= 1:
        last_col = get_column_letter(max_column)
        sheet.auto_filter.ref = f"A1:{last_col}{max_row}"
        sheet.freeze_panes = "A2"
    else:
        sheet.auto_filter.ref = None
        sheet.freeze_panes = None

    for column_index in range(1, max_column + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for row_index in range(1, min(max_row, 500) + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            max_length = max(max_length, len(str(value or "")))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 34)


def _payment_stage_for_excel(invoice: Invoice) -> str:
    if invoice.payment_stage == PaymentStage.PAID:
        return "Paid"
    if invoice.payment_stage == PaymentStage.FINANCE_REVIEW:
        return "Finance"
    return invoice.get_payment_stage_display()


def _team_label_from_invoice(invoice: Invoice) -> str:
    if invoice.team_id and invoice.team:
        return invoice.team.name
    return invoice.get_cost_bucket_display()


def _team_label_from_budget(line: BudgetLine) -> str:
    if line.team_id and line.team:
        return line.team.name
    return "No team"


def _raw(raw_data: dict, key: str, default):
    value = raw_data.get(key, default)
    if value is None:
        return default
    if isinstance(value, str) and not value.strip():
        return default
    return value


def _row_values(values: Iterable[Any]) -> list[Any]:
    return [_normalize_cell_value(value) for value in values]


def _normalize_cell_value(value: Any) -> Any:
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value != value or value in {float("inf"), float("-inf")}:  # NaN / inf
            return ""
        return value
    if isinstance(value, Decimal):
        return _excel_number(value)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None).isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return _sanitize_text(value)
    return _sanitize_text(str(value))


def _excel_number(value):
    if value in {"", None}:
        return ""
    try:
        decimal_value = Decimal(str(value).replace(",", ""))
    except Exception:
        return _sanitize_text(str(value))
    if not decimal_value.is_finite():
        return ""
    if decimal_value == decimal_value.to_integral_value():
        return int(decimal_value)
    return float(decimal_value)


# Control characters that are illegal in the Office Open XML format. openpyxl raises on these,
# and some can make spreadsheet apps misbehave, so we strip them defensively on every cell.
_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
# Excel's hard per-cell text limit.
_MAX_CELL_TEXT = 32000


def _sanitize_text(value):
    if not isinstance(value, str):
        return value
    cleaned = _ILLEGAL_XLSX_CHARS.sub("", value)
    if len(cleaned) > _MAX_CELL_TEXT:
        cleaned = cleaned[:_MAX_CELL_TEXT]
    return cleaned


def _list_value(values: list[str], index: int) -> str:
    if index >= len(values):
        return ""
    return values[index]

