from __future__ import annotations

from datetime import date
from decimal import Decimal

import pytest
from django.urls import reverse

from marketing.models import CostBucket, Invoice, PaymentStage

pytestmark = pytest.mark.django_db


def test_invoice_list_sort_by_amount(client, frontend_data):
    client.force_login(frontend_data["admin"])
    Invoice.objects.filter(pk=frontend_data["growth_invoice"].pk).update(amount=Decimal("1000"))
    Invoice.objects.filter(pk=frontend_data["brand_invoice"].pk).update(amount=Decimal("9000000"))
    response = client.get(reverse("marketing:invoice_list"), {"sort": "amount", "dir": "desc"})
    assert response.status_code == 200
    content = response.content.decode()
    brand_pos = content.find(str(frontend_data["brand_invoice"].invoice_number))
    growth_pos = content.find(str(frontend_data["growth_invoice"].invoice_number))
    assert brand_pos != -1 and growth_pos != -1
    assert brand_pos < growth_pos


def test_invoice_list_sortable_header_links(client, frontend_data):
    client.force_login(frontend_data["admin"])
    response = client.get(reverse("marketing:invoice_list"))
    assert response.status_code == 200
    assert 'class="sortable-th' in response.content.decode()
    assert "sort=amount" in response.content.decode()


def test_dashboard_renders_for_admin(client, frontend_data):
    client.force_login(frontend_data["admin"])

    response = client.get(reverse("marketing:dashboard"))

    assert response.status_code == 200
    assert "Finance overview" in response.content.decode()


def test_invoice_list_is_scoped_for_team_editor(client, frontend_data):
    client.force_login(frontend_data["growth_editor"])

    response = client.get(reverse("marketing:invoice_list"))
    content = response.content.decode()

    assert response.status_code == 200
    assert "G-UI-1" in content
    assert "B-UI-1" not in content


def test_observer_cannot_open_invoice_create(client, frontend_data):
    client.force_login(frontend_data["observer"])

    response = client.get(reverse("marketing:invoice_create"))

    assert response.status_code == 403


def test_editor_can_create_invoice_for_own_team(client, frontend_data):
    client.force_login(frontend_data["growth_editor"])

    response = client.post(
        reverse("marketing:invoice_create"),
        {
            "invoice_number": "G-UI-2",
            "vendor": frontend_data["vendor"].id,
            "team": frontend_data["growth"].id,
            "campaign": "",
            "category": "Performance",
            "cost_bucket": CostBucket.TEAM,
            "description": "",
            "invoice_date": "۱۴۰۵/۰۱/۱۴",
            "due_date": "",
            "amount": "1500.00",
            "currency": "IRR",
            "payment_stage": PaymentStage.SUBMITTED,
        },
    )

    assert response.status_code == 302
    invoice = Invoice.objects.get(invoice_number="G-UI-2", team=frontend_data["growth"])
    assert invoice.invoice_date == date(2026, 4, 3)


def test_invoice_list_filters_by_business_section(client, frontend_data):
    from django.urls import reverse

    invoice = frontend_data["growth_invoice"]
    invoice.business_section = "Consumer"
    invoice.save(update_fields=["business_section"])

    client.force_login(frontend_data["admin"])
    filtered = client.get(reverse("marketing:invoice_list"), {"business_section": "Consumer"})
    assert filtered.status_code == 200
    assert invoice in filtered.context["page_obj"]

    search = client.get(reverse("marketing:invoice_list"), {"q": "Consumer"})
    assert invoice in search.context["page_obj"]


def test_admin_can_export_invoice_excel_and_print_report(client, frontend_data):
    client.force_login(frontend_data["admin"])

    excel_response = client.get(reverse("marketing:export_invoices_excel"))
    print_response = client.get(reverse("marketing:invoice_report_print"))

    assert excel_response.status_code == 200
    assert excel_response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert print_response.status_code == 200
    assert "Invoice report" in print_response.content.decode()


def test_money_format_toggle_switches_display(client, frontend_data):
    client.force_login(frontend_data["admin"])
    Invoice.objects.filter(pk=frontend_data["brand_invoice"].pk).update(amount=Decimal("84276543010"))

    compact_response = client.get(reverse("marketing:dashboard"))
    compact_html = compact_response.content.decode()
    assert "84.3B" in compact_html or "84.3 میلیارد" in compact_html

    client.post(
        reverse("marketing:set_display_preferences"),
        {"money_format": "full", "next": reverse("marketing:dashboard")},
    )
    full_response = client.get(reverse("marketing:dashboard"))
    full_html = full_response.content.decode()
    assert "84,276,543,010" in full_html


def test_currency_unit_toggle_converts_to_toman(client, frontend_data):
    client.force_login(frontend_data["admin"])
    Invoice.objects.filter(pk=frontend_data["brand_invoice"].pk).update(amount=Decimal("84276543010"))

    client.post(
        reverse("marketing:set_display_preferences"),
        {"money_format": "full", "currency_unit": "toman", "next": reverse("marketing:dashboard")},
    )
    html = client.get(reverse("marketing:dashboard")).content.decode()

    assert "8,427,654,301" in html
    assert "84,276,543,010" not in html


def test_currency_unit_toggle_independent_of_compact_mode(client, frontend_data):
    client.force_login(frontend_data["admin"])
    Invoice.objects.filter(pk=frontend_data["brand_invoice"].pk).update(amount=Decimal("84276543010"))

    client.post(
        reverse("marketing:set_display_preferences"),
        {"money_format": "compact", "currency_unit": "toman", "next": reverse("marketing:dashboard")},
    )
    html = client.get(reverse("marketing:dashboard")).content.decode()

    assert "8.43B" in html


def test_workbook_style_export_recreates_source_sheets(client, frontend_data):
    from io import BytesIO

    from openpyxl import load_workbook

    from marketing.workbook_labels import DEFAULT_INVOICE_SHEET_NAME

    client.force_login(frontend_data["admin"])

    response = client.get(reverse("marketing:export_workbook_excel"))

    assert response.status_code == 200
    assert response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == [
        DEFAULT_INVOICE_SHEET_NAME,
        "Budget",
        "Market Live Spending",
        "Data",
    ]
    invoice_sheet = workbook[DEFAULT_INVOICE_SHEET_NAME]
    header = [cell.value for cell in invoice_sheet[1]]
    assert "Invoice Number" in header
    assert "Invoice Amount (IRR)" in header
    invoice_number_column = header.index("Invoice Number") + 1
    cell_values = {row[invoice_number_column - 1].value for row in invoice_sheet.iter_rows(min_row=2)}
    assert "G-UI-1" in cell_values


def test_workbook_style_export_blocked_for_observer(client, frontend_data):
    client.force_login(frontend_data["observer"])

    response = client.get(reverse("marketing:export_workbook_excel"))

    assert response.status_code == 403


def test_theme_toggle_sets_dark_body_class(client, frontend_data):
    client.force_login(frontend_data["admin"])

    default_html = client.get(reverse("marketing:dashboard")).content.decode()
    assert "theme-light" in default_html

    client.post(
        reverse("marketing:set_display_preferences"),
        {"theme": "dark", "next": reverse("marketing:dashboard")},
    )
    dark_html = client.get(reverse("marketing:dashboard")).content.decode()
    assert "theme-dark" in dark_html


def test_dashboard_chart_helpers_load_before_chart_init(client, frontend_data):
    client.force_login(frontend_data["admin"])

    html = client.get(reverse("marketing:dashboard")).content.decode()

    helper_pos = html.find("window.chartMoneyOptions = function")
    variance_helper_pos = html.find("window.createBudgetVarianceChart = function")
    init_pos = html.find("new Chart(monthlyEl")
    assert helper_pos != -1
    assert variance_helper_pos != -1
    assert init_pos != -1
    assert helper_pos < init_pos, "Chart money helpers must load before dashboard chart init"
    assert variance_helper_pos < init_pos, "Budget variance helper must load before dashboard chart init"
    assert "yAxisID: \"y\"" in html or "createBudgetVarianceChart" in html
