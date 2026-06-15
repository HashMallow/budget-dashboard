from __future__ import annotations

import shutil
import subprocess
import zipfile
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from django.urls import reverse
from openpyxl import load_workbook

from marketing.exports.workbook import _normalize_cell_value, build_workbook_style_export
from marketing.models import BudgetLine, Invoice
from marketing.workbook_labels import DEFAULT_INVOICE_SHEET_NAME


def _allowed_cell_types(value) -> bool:
    return value is None or isinstance(value, (str, int, float, bool))


@pytest.mark.django_db
def test_workbook_cells_use_excel_safe_types(frontend_data):
    workbook = build_workbook_style_export(Invoice.objects.all(), BudgetLine.objects.all())
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    loaded = load_workbook(buffer)
    for sheet in loaded.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                assert _allowed_cell_types(cell.value), (
                    f"{sheet.title}!{cell.coordinate} has unsupported type {type(cell.value)}"
                )
                if isinstance(cell.value, str):
                    assert "\x00" not in cell.value
                    assert len(cell.value) <= 32000


@pytest.mark.django_db
def test_workbook_xlsx_is_valid_zip_package(frontend_data):
    workbook = build_workbook_style_export(Invoice.objects.all(), BudgetLine.objects.all())
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    assert zipfile.is_zipfile(buffer)


@pytest.mark.django_db
def test_workbook_auto_filter_matches_data_bounds(frontend_data):
    workbook = build_workbook_style_export(Invoice.objects.all(), BudgetLine.objects.all())
    buffer = BytesIO()
    workbook.save(buffer)
    loaded = load_workbook(BytesIO(buffer.getvalue()))
    invoice_sheet = loaded[DEFAULT_INVOICE_SHEET_NAME]
    assert invoice_sheet.auto_filter.ref == f"A1:Y{invoice_sheet.max_row}"


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (None, ""),
        ("hello", "hello"),
        (Decimal("1250000"), 1250000),
        (float("nan"), ""),
        (date(2025, 3, 1), "2025-03-01"),
        (datetime(2025, 3, 1, 12, 30), "2025-03-01 12:30:00"),
        ({"bad": "object"}, "object"),
    ],
)
def test_normalize_cell_value_coerces_to_safe_scalars(value, expected):
    result = _normalize_cell_value(value)
    if isinstance(value, dict):
        assert expected in result
    else:
        assert result == expected


@pytest.mark.django_db
def test_workbook_export_view_roundtrip(client, frontend_data):
    client.force_login(frontend_data["admin"])
    response = client.get(reverse("marketing:export_workbook_excel"))
    assert response.status_code == 200
    load_workbook(BytesIO(response.content))


@pytest.mark.django_db
@pytest.mark.skipif(not shutil.which("libreoffice"), reason="LibreOffice not installed")
def test_workbook_converts_with_libreoffice(tmp_path, frontend_data):
    workbook = build_workbook_style_export(Invoice.objects.all(), BudgetLine.objects.all())
    source = tmp_path / "marketing-workbook.xlsx"
    out_dir = tmp_path / "converted"
    out_dir.mkdir()
    workbook.save(source)

    result = subprocess.run(
        [
            shutil.which("libreoffice") or "libreoffice",
            "--headless",
            "--convert-to",
            "csv",
            "--outdir",
            str(out_dir),
            str(source),
        ],
        capture_output=True,
        text=True,
        timeout=60,
        check=False,
    )
    assert result.returncode == 0, result.stderr or result.stdout
    if not any(out_dir.glob("*.csv")):
        pytest.skip("LibreOffice ran but wrote no CSV (common with Snap sandbox + /tmp paths)")
