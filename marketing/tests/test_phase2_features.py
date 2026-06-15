from __future__ import annotations

from datetime import date
from decimal import Decimal

import pytest
import yaml
from django.urls import reverse
from openpyxl import Workbook

from marketing.models import (
    CostBucket,
    Invoice,
    PaymentStage,
    Requester,
    Role,
    SpendCategory,
    SubTeam,
    Team,
    UserTeamAccess,
    Vendor,
)
from marketing.reference_data import load_workbook_data, seed_reference_data_from_workbook
from marketing.tests.test_excel_importer import write_mapping, write_workbook

pytestmark = pytest.mark.django_db


def _write_reference_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(
        [
            "requester",
            "Vendor List",
            "Vendor list unique",
            "Title",
            "Title list unique",
            "Month list",
            "sub team",
        ]
    )
    sheet.append(["Ali", "Vendor A", "Vendor A", "Performance", "Performance", "Farvardin", "Growth Ops"])
    sheet.append(["Sara", "Vendor B", "", "Brand", "", "Ordibehesht", ""])
    workbook.save(path)


@pytest.fixture
def reference_workbook(tmp_path):
    path = tmp_path / "reference.xlsx"
    _write_reference_workbook(path)
    return path


def _append_data_sheet(workbook_path):
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path)
    data = workbook.create_sheet("Data")
    data.append(
        [
            "requester",
            "Vendor List",
            "Vendor list unique",
            "Title",
            "Title list unique",
            "Month list",
            "sub team",
        ]
    )
    data.append(["Ali", "Lookup Vendor", "Lookup Vendor", "Performance", "Performance", "Farvardin", "Growth Ops"])
    workbook.save(workbook_path)


@pytest.fixture
def full_workbook(tmp_path):
    workbook_path = tmp_path / "full.xlsx"
    mapping_path = tmp_path / "column_mapping.yml"
    write_workbook(workbook_path)
    _append_data_sheet(workbook_path)
    write_mapping(mapping_path, workbook_path)
    mapping = yaml.safe_load(mapping_path.read_text(encoding="utf-8"))
    mapping["sheets"]["lookup_data"] = {
        "actual_sheet_name": "Data",
        "header_row": 1,
        "columns": {
            "requester_name": "requester",
            "vendor_name": "Vendor List",
            "unique_vendor_name": "Vendor list unique",
            "title": "Title",
            "unique_title": "Title list unique",
            "sub_team": "sub team",
        },
    }
    mapping_path.write_text(yaml.safe_dump(mapping, allow_unicode=True), encoding="utf-8")
    return workbook_path, mapping_path


def test_seed_reference_data_creates_lookup_rows(reference_workbook):
    result = seed_reference_data_from_workbook(reference_workbook, dry_run=False)

    assert Vendor.objects.filter(name="Vendor A").exists()
    assert SpendCategory.objects.filter(name="Performance").exists()
    assert SubTeam.objects.filter(name="Growth Ops").exists()
    assert Requester.objects.filter(name="Ali").exists()
    assert result.vendors.created >= 1
    assert result.categories.created >= 1


def test_load_workbook_data_seeds_reference_lookups(full_workbook):
    """UI import and make load-data both seed Data-sheet lookups, not just invoices."""
    workbook_path, mapping_path = full_workbook
    load_result = load_workbook_data(workbook_path, mapping_path=mapping_path, dry_run=False)

    assert SpendCategory.objects.filter(name="Performance").exists()
    assert SubTeam.objects.filter(name="Growth Ops").exists()
    assert Requester.objects.filter(name="Ali").exists()
    assert load_result.reference_result.categories.created >= 1
    assert load_result.reference_result.requesters.created >= 1


def test_import_workbook_view_seeds_reference_lookups(client, full_workbook):
    from django.contrib.auth import get_user_model

    workbook_path, _mapping_path = full_workbook
    admin = get_user_model().objects.create_superuser(username="import-admin", password="test-pass")
    client.force_login(admin)

    with workbook_path.open("rb") as workbook_file:
        response = client.post(
            reverse("marketing:import_workbook"),
            {"action": "dry_run", "workbook": workbook_file},
        )
    assert response.status_code == 200

    confirm = client.post(reverse("marketing:import_workbook"), {"action": "confirm"})
    assert confirm.status_code == 200

    assert SpendCategory.objects.filter(name="Performance").exists()
    assert Requester.objects.filter(name="Ali").exists()


def test_team_dashboard_is_scoped_for_editor(client):
    from django.contrib.auth import get_user_model

    user_model = get_user_model()
    editor = user_model.objects.create_user(username="editor", password="pass")
    growth = Team.objects.create(name="Growth", slug="growth")
    brand = Team.objects.create(name="Brand", slug="brand")
    vendor = Vendor.objects.create(name="Vendor")
    UserTeamAccess.objects.create(user=editor, team=growth, role=Role.EDITOR)
    Invoice.objects.create(
        invoice_number="G-1",
        vendor=vendor,
        team=growth,
        category="Performance",
        cost_bucket=CostBucket.TEAM,
        invoice_date=date(2026, 4, 1),
        amount=Decimal("1000"),
        payment_stage=PaymentStage.PAID,
    )
    Invoice.objects.create(
        invoice_number="B-1",
        vendor=vendor,
        team=brand,
        category="Brand",
        cost_bucket=CostBucket.TEAM,
        invoice_date=date(2026, 4, 2),
        amount=Decimal("2000"),
        payment_stage=PaymentStage.PAID,
    )

    client.force_login(editor)
    ok = client.get(reverse("marketing:team_dashboard", args=[growth.pk]))
    forbidden = client.get(reverse("marketing:team_dashboard", args=[brand.pk]))

    assert ok.status_code == 200
    assert "G-1" in ok.content.decode()
    assert forbidden.status_code == 404


def test_admin_can_export_vendor_campaign_excel_and_pdf(client):
    from django.contrib.auth import get_user_model

    admin = get_user_model().objects.create_superuser(username="admin", password="pass")
    team = Team.objects.create(name="Growth", slug="growth")
    vendor = Vendor.objects.create(name="Vendor")
    Invoice.objects.create(
        invoice_number="1",
        vendor=vendor,
        team=team,
        category="Performance",
        cost_bucket=CostBucket.TEAM,
        invoice_date=date(2026, 4, 1),
        amount=Decimal("1000"),
        payment_stage=PaymentStage.PAID,
    )

    client.force_login(admin)
    vendor_xlsx = client.get(reverse("marketing:export_vendors_excel"))
    campaign_xlsx = client.get(reverse("marketing:export_campaigns_excel"))
    pdf = client.get(reverse("marketing:dashboard_report_pdf"))

    assert vendor_xlsx.status_code == 200
    assert vendor_xlsx["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert campaign_xlsx.status_code == 200
    assert pdf.status_code == 200
    assert pdf["Content-Type"] == "application/pdf"
    assert pdf.content.startswith(b"%PDF")
