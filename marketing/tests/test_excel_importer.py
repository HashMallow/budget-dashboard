from __future__ import annotations

from datetime import date

import pytest
import yaml
from openpyxl import Workbook

from marketing.importers.excel import import_marketing_workbook, parse_excel_date, resolve_sheet_name
from marketing.models import BudgetLine, Campaign, CostBucket, Invoice, PaymentStage, Team, TeamAlias, Vendor

pytestmark = pytest.mark.django_db


def test_resolve_sheet_name_auto_detects_by_headers_when_tab_name_differs():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Real Customer Tab Name"
    sheet.append(
        [
            "Year",
            "MKT Team",
            "Budget Line",
            "Campaign Name",
            "Vendor Name",
            "Description",
            "invoice date in gregorian",
            "invoice date in Jalali",
            "Invoice Number",
            "Invoice Amount (IRR)",
            "payment state",
            "MKT to Finance sent date in gregorian",
            "MKT to Finance sent date",
            "blank_AC",
        ]
    )
    invoice_mapping = {
        "actual_sheet_name": "Marketing Spend Input",
        "header_row": 1,
        "columns": {
            "invoice_number": "Invoice Number",
            "vendor_name": "Vendor Name",
            "amount": "Invoice Amount (IRR)",
        },
    }
    assert resolve_sheet_name(workbook, invoice_mapping, invoice=True) == "Real Customer Tab Name"


def test_jalali_text_dates_are_parsed_as_shamsi_before_gregorian():
    assert parse_excel_date("1405/01/10", None) == date(2026, 3, 30)
    assert parse_excel_date("۱۴۰۵/۰۱/۱۰", None) == date(2026, 3, 30)


def write_mapping(path, workbook_path):
    mapping = {
        "workbook": str(workbook_path),
        "sheets": {
            "invoices": {
                "actual_sheet_name": "Invoices",
                "header_row": 1,
                "data_start_row": 2,
                "data_end_row": 3,
                "columns": {
                    "year": "Year",
                    "team": "MKT Team",
                    "category": "Budget Line",
                    "campaign_name": "Campaign Name",
                    "vendor_name": "Vendor Name",
                    "description": "Description",
                    "invoice_date_gregorian_serial": "invoice date in gregorian",
                    "invoice_date_jalali_serial": "invoice date in Jalali",
                    "invoice_number": "Invoice Number",
                    "amount": "Invoice Amount (IRR)",
                    "payment_stage": "payment state",
                    "finance_sent_date_gregorian_serial": "MKT to Finance sent date in gregorian",
                    "finance_sent_date_jalali_serial": "MKT to Finance sent date",
                    "jalali_invoice_date_text_candidate": "blank_AC",
                },
                "derived_values": {
                    "currency": "IRR",
                    "cost_bucket": {
                        "default": "TEAM",
                        "rules": [
                            {
                                "bucket": "REFERRAL",
                                "when_any_source_column_contains": ["referral", "ریفرال"],
                                "source_columns": ["MKT Team", "Budget Line", "Vendor Name", "Description"],
                            }
                        ],
                    },
                    "payment_stage_mapping": {"Paid": "PAID", "Finance": "FINANCE_REVIEW"},
                },
            },
            "budget": {
                "actual_sheet_name": "Budget",
                "header_row": 1,
                "data_start_row": 2,
                "data_end_row": 2,
                "row_context_columns": {
                    "team": "Team",
                    "category_or_title": "Title",
                    "description": "Description",
                    "currency": "Unit Price",
                },
                "budget_line_mapping": {"year": 1405},
                "monthly_columns": [
                    {
                        "month_number": 1,
                        "month_label": "Farvardin",
                        "projection_column": "E",
                        "actual_column": "F",
                    },
                    {
                        "month_number": 2,
                        "month_label": "Ordibehesht",
                        "projection_column": "G",
                        "actual_column": "H",
                    },
                ],
            },
        },
    }
    path.write_text(yaml.safe_dump(mapping, allow_unicode=True), encoding="utf-8")


def write_workbook(path):
    workbook = Workbook()
    invoices = workbook.active
    invoices.title = "Invoices"
    invoices.append(
        [
            "Year",
            "MKT Team",
            "Budget Line",
            "Campaign Name",
            "Vendor Name",
            "Description",
            "invoice date in gregorian",
            "invoice date in Jalali",
            "Invoice Number",
            "Invoice Amount (IRR)",
            "payment state",
            None,
            None,
            None,
            "blank_AC",
            "MKT to Finance sent date in gregorian",
            "MKT to Finance sent date",
        ]
    )
    invoices.append(
        [
            1405,
            "Growth",
            "Performance",
            "on going",
            "Test Vendor",
            "Regular invoice",
            date(2026, 3, 30),
            None,
            "1",
            1000,
            "Paid",
            None,
            None,
            None,
            "1405/01/10",
            None,
            None,
        ]
    )
    invoices.append(
        [
            1405,
            "Referral",
            "referral",
            "on going",
            "Test Vendor",
            "Referral invoice",
            date(2026, 4, 20),
            None,
            "1",
            2000,
            "Finance",
            None,
            None,
            None,
            "1405/01/31",
            date(2026, 5, 15),
            None,
        ]
    )

    budget = workbook.create_sheet("Budget")
    budget.append(["Team", "Title", "Description", "Unit Price", "projection", "Actual", "projection", "Actual"])
    budget.append(["Growth", "Performance", "test budget", "Rial", 3000, 1000, 4000, 0])
    workbook.save(path)


def test_import_workbook_creates_invoices_budget_lines_and_preserves_duplicates(tmp_path):
    workbook_path = tmp_path / "marketing.xlsx"
    mapping_path = tmp_path / "column_mapping.yml"
    write_workbook(workbook_path)
    write_mapping(mapping_path, workbook_path)

    result = import_marketing_workbook(workbook_path, mapping_path=mapping_path)

    assert result.invoices.created == 2
    assert result.budget_lines.created == 2
    assert Invoice.objects.count() == 2
    assert Vendor.objects.count() == 1
    assert Team.objects.filter(name="Growth").exists()
    assert not Team.objects.filter(name="Referral", is_active=True).exists()
    referral = Invoice.objects.get(cost_bucket=CostBucket.REFERRAL)
    assert referral.team.name == "Growth"
    assert Invoice.objects.filter(invoice_number="1", vendor__name="Test Vendor").count() == 2
    assert Invoice.objects.filter(cost_bucket=CostBucket.REFERRAL).count() == 1
    assert Invoice.objects.filter(payment_stage=PaymentStage.PAID).count() == 1
    assert Invoice.objects.filter(payment_stage=PaymentStage.FINANCE_REVIEW).count() == 1
    assert BudgetLine.objects.count() == 2
    # Free-text "on going" is a workbook placeholder, not a real campaign.
    assert Campaign.objects.count() == 0
    assert Invoice.objects.filter(campaign__isnull=True).count() == 2

    second_result = import_marketing_workbook(workbook_path, mapping_path=mapping_path)

    assert second_result.invoices.created == 0
    assert second_result.invoices.updated == 2
    assert second_result.budget_lines.created == 0
    assert second_result.budget_lines.updated == 2
    assert Invoice.objects.count() == 2
    assert BudgetLine.objects.count() == 2


def test_dry_run_does_not_write_database(tmp_path):
    workbook_path = tmp_path / "marketing.xlsx"
    mapping_path = tmp_path / "column_mapping.yml"
    write_workbook(workbook_path)
    write_mapping(mapping_path, workbook_path)

    result = import_marketing_workbook(workbook_path, mapping_path=mapping_path, dry_run=True)

    assert result.invoices.created == 2
    assert result.budget_lines.created == 2
    assert Invoice.objects.count() == 0
    assert BudgetLine.objects.count() == 0
    assert Vendor.objects.count() == 0


def test_canonical_team_name_folds_workbook_variants():
    from marketing.importers.excel import canonical_team_name

    assert canonical_team_name("Operation & Analysis") == "Ops & Analytics"
    assert canonical_team_name("Brand (PR & Social & CSR)") == "Brand"
    assert canonical_team_name("Growth") == "Growth"


def test_import_merges_workbook_team_spelling_variants(tmp_path):
    """Existing duplicate variants fold into the canonical team on re-import."""
    workbook_path = tmp_path / "marketing.xlsx"
    mapping_path = tmp_path / "column_mapping.yml"
    write_workbook(workbook_path)
    write_mapping(mapping_path, workbook_path)

    canonical = Team.objects.create(name="Ops & Analytics", slug="ops-analytics")
    variant = Team.objects.create(name="Operation & Analysis", slug="operation-analysis")
    vendor = Vendor.objects.create(name="Variant Vendor")
    Invoice.objects.create(
        invoice_number="OA-1",
        vendor=vendor,
        team=variant,
        category="Performance",
        cost_bucket=CostBucket.TEAM,
        invoice_date=date(2026, 3, 30),
        amount=500,
    )

    import_marketing_workbook(workbook_path, mapping_path=mapping_path)

    canonical.refresh_from_db()
    variant.refresh_from_db()
    assert canonical.is_active
    assert variant.is_active is False
    assert Invoice.objects.get(invoice_number="OA-1").team == canonical
    assert TeamAlias.objects.filter(raw_name="Operation & Analysis", team=canonical).exists()


def test_finance_review_invoice_uses_finance_sent_date_for_stage_aging(tmp_path):
    workbook_path = tmp_path / "marketing.xlsx"
    mapping_path = tmp_path / "column_mapping.yml"
    write_workbook(workbook_path)
    write_mapping(mapping_path, workbook_path)

    import_marketing_workbook(workbook_path, mapping_path=mapping_path)

    finance_invoice = Invoice.objects.get(payment_stage=PaymentStage.FINANCE_REVIEW)
    # Aging must come from the workbook's "MKT to Finance sent date", not the import timestamp.
    assert finance_invoice.stage_changed_at.date() == date(2026, 5, 15)
    paid_invoice = Invoice.objects.get(payment_stage=PaymentStage.PAID)
    assert paid_invoice.stage_changed_at.date() != date(2026, 5, 15)


def test_budget_reimport_is_idempotent_on_source_row_even_if_team_changes(tmp_path):
    workbook_path = tmp_path / "marketing.xlsx"
    mapping_path = tmp_path / "column_mapping.yml"
    write_workbook(workbook_path)
    write_mapping(mapping_path, workbook_path)

    # A stale row from an earlier import with a different team/category for the same source cell.
    BudgetLine.objects.create(
        year=1405,
        month=1,
        team=None,
        category="Stale category",
        planned_amount=1,
        currency="IRR",
        source_sheet="Budget",
        source_row_number=2,
    )

    import_marketing_workbook(workbook_path, mapping_path=mapping_path)

    rows = BudgetLine.objects.filter(source_sheet="Budget", source_row_number=2, year=1405, month=1)
    assert rows.count() == 1
    # The stale row was updated in place, not duplicated.
    assert rows.first().category == "Performance"
    assert rows.first().team.name == "Growth"


def test_import_resolves_team_aliases(tmp_path):
    workbook_path = tmp_path / "marketing.xlsx"
    mapping_path = tmp_path / "column_mapping.yml"
    write_workbook(workbook_path)
    write_mapping(mapping_path, workbook_path)

    canonical = Team.objects.create(name="Ops & Analytics", slug="ops-analytics")
    TeamAlias.objects.create(raw_name="Growth", team=canonical)

    import_marketing_workbook(workbook_path, mapping_path=mapping_path)

    invoice = Invoice.objects.get(invoice_number="1", cost_bucket=CostBucket.TEAM)
    assert invoice.team == canonical
    assert not Team.objects.filter(name="Growth").exists()
